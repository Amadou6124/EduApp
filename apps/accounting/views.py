"""Vues Comptabilité — paie, dépenses, bilan (Phases 2-7)."""
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

from apps.accounts.models import UserRole
from django.views.decorators.http import require_http_methods

from apps.core.mixins import (
    get_school, director_or_accounting_required, director_or_emargement_required,
)


def _toast_error(message):
    resp = HttpResponse(status=422)
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': message, 'type': 'error'}})
    return resp


def _parse_money(raw):
    raw = (raw or '').strip().replace(' ', '')
    if not raw:
        return None
    try:
        v = Decimal(raw)
    except InvalidOperation:
        return None
    return v if v >= 0 else None


# ─── Phase 2 — Liste employés & rémunération ─────────────────────────────────

@login_required
@director_or_accounting_required
def accounting_staff_list(request):
    """Liste des membres payables de l'école + leur profil de rémunération. Zéro N+1."""
    from apps.accounts.models import Membership

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    memberships = (
        Membership.objects
        .filter(school=school, is_active=True)
        .exclude(role__in=[UserRole.PARENT, UserRole.STUDENT])
        .select_related('user', 'employee_profile')   # reverse OneToOne → 1 requête
        .order_by('role', 'user__full_name')
    )

    items = []
    counts = {'all': 0, 'permanent': 0, 'vacataire': 0, 'none': 0}
    for m in memberships:
        profile = getattr(m, 'employee_profile', None)   # RelatedObjectDoesNotExist ⊂ AttributeError → None
        if profile is None:
            cat = 'none'
        elif profile.employment_type == 'permanent':
            cat = 'permanent'
        else:
            cat = 'vacataire'
        items.append({'membership': m, 'profile': profile, 'cat': cat})
        counts['all'] += 1
        counts[cat] += 1

    return render(request, 'accounting/staff_list.html', {
        'items': items, 'counts': counts, 'school': school,
        'no_profile_count': counts['none'],
    })


# ─── Phase 2 — Rémunération employé (panneau lazy-load dans /team/<id>/) ──────

def _vacataire_course_rates(school, member, profile):
    """Cours du prof (matière + classe) groupés par matière + leur tarif actuel.

    Auto-listés depuis ses cours → tarif par cours, avec « appliquer à toute la
    matière » côté UI pour éviter la re-saisie.
    """
    from apps.schools.models import ClassSubject
    from .models import VacataireRate

    courses = (
        ClassSubject.objects
        .filter(teacher=member, school_class__school=school, is_active=True)
        .select_related('subject', 'school_class')
        .order_by('subject__name', 'school_class__name')
    )
    existing = {
        vr.class_subject_id: vr.hourly_rate
        for vr in VacataireRate.objects.filter(profile=profile)
    }
    groups, by_subj = [], {}
    for cs in courses:
        g = by_subj.get(cs.subject_id)
        if g is None:
            g = {'subject': cs.subject, 'courses': []}
            by_subj[cs.subject_id] = g
            groups.append(g)
        g['courses'].append({'cs': cs, 'rate': existing.get(cs.id)})
    return groups


@login_required
@director_or_accounting_required
def employee_remuneration_panel(request, user_id):
    """Panneau rémunération (GET, lazy-load HTMX). Module désactivé → état clair
    en 200 (jamais un 403 nu : HTMX ne remplacerait rien → « Chargement… » figé)."""
    from apps.accounts.models import User, Membership
    from .models import EmployeeProfile, EmploymentType

    school = get_school(request)
    if not school.accounting_enabled:
        return render(request, 'accounting/partials/remuneration_disabled.html',
                      {'school': school})

    member = get_object_or_404(User, pk=user_id, memberships__school=school)
    membership = get_object_or_404(Membership, user=member, school=school)
    profile, _ = EmployeeProfile.objects.get_or_create(
        membership=membership,
        defaults={'employment_type': EmploymentType.PERMANENT},
    )
    return render(request, 'accounting/partials/employee_remuneration.html', {
        'member': member, 'profile': profile, 'school': school,
        'rate_groups': _vacataire_course_rates(school, member, profile),
    })


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def employee_remuneration_save(request, user_id):
    """Sauvegarde EmployeeProfile. Permanent → salaire requis ; vacataire → taux requis."""
    from apps.accounts.models import User, Membership
    from .models import EmployeeProfile, EmploymentType

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    member = get_object_or_404(User, pk=user_id, memberships__school=school)
    membership = get_object_or_404(Membership, user=member, school=school)
    profile, _ = EmployeeProfile.objects.get_or_create(membership=membership)

    emp_type = request.POST.get('employment_type', 'permanent')
    if emp_type not in (EmploymentType.PERMANENT, EmploymentType.VACATAIRE):
        emp_type = EmploymentType.PERMANENT

    monthly = _parse_money(request.POST.get('monthly_salary'))

    if emp_type == EmploymentType.PERMANENT and monthly is None:
        return _toast_error('Le salaire mensuel est obligatoire pour un permanent.')

    hire_raw = (request.POST.get('hire_date') or '').strip()
    hire_date = None
    if hire_raw:
        try:
            hire_date = datetime.strptime(hire_raw, '%Y-%m-%d').date()
        except ValueError:
            hire_date = None

    profile.employment_type = emp_type
    profile.monthly_salary = monthly if emp_type == EmploymentType.PERMANENT else None
    profile.hourly_rate    = None  # remplacé par les tarifs par matière (VacataireRate)
    profile.hire_date = hire_date
    profile.save()

    # Tarifs par cours (vacataire) : upsert des valeurs saisies, suppression si vidé.
    if emp_type == EmploymentType.VACATAIRE:
        from .models import VacataireRate
        for g in _vacataire_course_rates(school, member, profile):
            for r in g['courses']:
                cs_id = r['cs'].id
                raw = (request.POST.get(f'rate_cs_{cs_id}') or '').strip()
                if raw == '':
                    VacataireRate.objects.filter(profile=profile, class_subject_id=cs_id).delete()
                    continue
                val = _parse_money(raw)
                if val is not None and val >= 0:
                    VacataireRate.objects.update_or_create(
                        profile=profile, class_subject_id=cs_id, defaults={'hourly_rate': val},
                    )

    resp = render(request, 'accounting/partials/employee_remuneration.html', {
        'member': member, 'profile': profile, 'school': school,
        'rate_groups': _vacataire_course_rates(school, member, profile),
    })
    resp['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Rémunération enregistrée.', 'type': 'success'},
    })
    return resp


# ─── Phase 3 — Émargement enseignants ────────────────────────────────────────

_LEVEL_BADGE = {
    'prescolaire':    'bg-purple-100 text-purple-700',
    'fondamental_1':  'bg-blue-100 text-blue-700',
    'fondamental_2':  'bg-indigo-100 text-indigo-700',
    'secondaire_gen': 'bg-green-100 text-green-700',
    'secondaire_pro': 'bg-teal-100 text-teal-700',
    'superieur':      'bg-orange-100 text-orange-700',
}
_SESSIONS = ['morning', 'afternoon', 'full']


def _hour_presets(duration):
    """4 raccourcis d'heures se terminant sur la durée prévue (ex. 2h → 0,5/1/1,5/2 ;
    3h → 1,5/2/2,5/3). Le « partiel » exact reste possible en saisie libre."""
    d = float(duration or 2)
    vals = [round(d - 1.5, 1), round(d - 1.0, 1), round(d - 0.5, 1), round(d, 1)]
    out = []
    for v in vals:
        if v >= 0.5:
            s = ('%g' % v)  # 1.5 / 2
            out.append({'v': s, 'label': s.replace('.', ',')})
    return out


@login_required
@director_or_emargement_required
def emargement_dashboard(request):
    """Émargement « carnet » : séances vacataires à émarger (heures réelles +
    tarif) + permanents présumés présents (on signale les exceptions). Honnête :
    pas de planning → on enregistre ce qui a eu lieu, le résumé compte le réel."""
    from datetime import datetime as _dt, date as _date, timedelta
    from collections import OrderedDict
    from apps.schools.models import ClassSubject
    from .models import TeacherAttendance, EmployeeProfile, EmploymentType, VacataireRate

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    today = _date.today()
    try:
        selected_date = _dt.strptime(request.GET.get('date', ''), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        selected_date = today

    # Qui est vacataire (école, actif) → quels cours sont « à émarger ».
    vac_user_ids = set(
        EmployeeProfile.objects.filter(
            membership__school=school, membership__is_active=True,
            employment_type=EmploymentType.VACATAIRE, is_active=True,
        ).values_list('membership__user_id', flat=True)
    )

    cs_list = list(
        ClassSubject.objects
        .filter(school_class__school=school, school_class__is_active=True,
                is_active=True, teacher__isnull=False)
        .select_related('subject', 'teacher', 'school_class')
        .order_by('teacher__full_name', 'subject__name', 'school_class__name')
    )
    vac_ids = [cs.id for cs in cs_list if cs.teacher_id in vac_user_ids]
    rates = {
        vr.class_subject_id: vr.hourly_rate
        for vr in VacataireRate.objects.filter(class_subject_id__in=vac_ids)
    }

    # ── Pré-filtrage par l'emploi du temps (GUIDE, jamais verrou) ──────────────
    # Les cours ayant un créneau CE jour-là remontent en tête avec leur horaire ;
    # tous les autres restent émargeables (cours exceptionnel = réalité du terrain).
    # Python weekday() : lundi=0 … dimanche=6 — même convention que Weekday.
    from apps.schools.models import CourseSlot
    from apps.schools.periods import active_year_for
    year = active_year_for(school)
    planned_map, edt_in_use = {}, False
    if year:
        slot_qs = CourseSlot.objects.filter(
            school_year=year, class_subject__school_class__school=school,
        )
        edt_in_use = slot_qs.exists()
        for s in slot_qs.filter(day=selected_date.weekday()).order_by('start_time'):
            planned_map.setdefault(s.class_subject_id, []).append(
                f'{s.start_time:%H:%M}–{s.end_time:%H:%M}'
            )

    att_map = {
        a.class_subject_id: a
        for a in TeacherAttendance.objects
        .filter(school=school, date=selected_date, session='morning')
        .select_related('substitute')
    }

    init_state, vac_groups, perm_by_class = {}, OrderedDict(), OrderedDict()
    sessions_done, day_amount, absences = 0, 0, 0
    perm_teacher_ids = set()

    for cs in cs_list:
        a = att_map.get(cs.id)
        status = a.status if a else ''
        actual = a.hours if (a and a.hours is not None) else None
        init_state[str(cs.id)] = {
            'status':   status,
            'hours':    (str(actual) if actual is not None else ''),
            'sub_id':   str(a.substitute_id) if a and a.substitute_id else '',
            'sub_name': a.substitute.full_name if a and a.substitute else '',
        }
        if status == 'absent':
            absences += 1

        if cs.teacher_id in vac_user_ids:
            rate = rates.get(cs.id)
            eff = actual if actual is not None else cs.duration_hours
            if status in ('present', 'replaced'):
                sessions_done += 1
            if status == 'present' and rate:
                day_amount += rate * eff
            g = vac_groups.setdefault(cs.teacher_id, {'teacher': cs.teacher, 'courses': []})
            g['courses'].append({
                'cs': cs, 'rate': rate, 'duration': cs.duration_hours,
                'presets': _hour_presets(cs.duration_hours),
                'planned': planned_map.get(cs.id),   # horaires prévus aujourd'hui (ou None)
            })
        else:
            perm_teacher_ids.add(cs.teacher_id)
            g = perm_by_class.setdefault(cs.school_class_id, {'sc': cs.school_class, 'courses': []})
            g['courses'].append({'cs': cs})

    # Tri « prévus d'abord » : dans chaque groupe, les cours planifiés du jour en tête
    # (par heure de début) ; puis les groupes ayant ≥1 cours prévu avant les autres.
    for g in vac_groups.values():
        g['courses'].sort(key=lambda c: (c['planned'] is None,
                                         c['planned'][0] if c['planned'] else ''))
    vac_groups = [{
        'teacher': g['teacher'], 'courses': g['courses'],
        'course_ids': [c['cs'].id for c in g['courses']],
        'planned_count': sum(1 for c in g['courses'] if c['planned']),
    } for g in vac_groups.values()]
    vac_groups.sort(key=lambda g: g['planned_count'] == 0)
    perm_groups = []
    for e in perm_by_class.values():
        teachers = ' '.join(sorted({c['cs'].teacher.full_name for c in e['courses']}))
        perm_groups.append({
            'school_class': e['sc'], 'courses': e['courses'],
            'search': f"{e['sc'].name} {teachers}".lower(),
        })

    return render(request, 'accounting/emargement.html', {
        'school': school,
        'selected_date': selected_date,
        'today': today,
        'prev_date': selected_date - timedelta(days=1),
        'next_date': selected_date + timedelta(days=1),
        'vac_groups': vac_groups,
        'vac_teacher_count': len(vac_groups),
        'edt_in_use': edt_in_use,
        'planned_total': sum(g['planned_count'] for g in vac_groups),
        'perm_groups': perm_groups,
        'perm_total': sum(len(g['courses']) for g in perm_groups),
        'perm_teacher_count': len(perm_teacher_ids),
        'summary': {'sessions_done': sessions_done, 'day_amount': day_amount, 'absences': absences},
        'init_state': init_state,
    })


@login_required
@director_or_emargement_required
@require_http_methods(['POST'])
def emargement_save(request):
    """Persiste un émargement (Alpine gère l'UI → 204). Anti-fraude : user ≠ teacher du cours."""
    from datetime import datetime as _dt
    from apps.schools.models import ClassSubject
    from apps.accounts.models import User
    from .models import TeacherAttendance, SessionType, TeacherAttendanceStatus

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    try:
        cs_id = int(request.POST.get('class_subject_id'))
    except (TypeError, ValueError):
        return HttpResponse(status=400)
    cs = get_object_or_404(
        ClassSubject, pk=cs_id, school_class__school=school, is_active=True, teacher__isnull=False,
    )

    # Anti-fraude : un enseignant ne peut pas émarger son propre cours
    if cs.teacher_id == request.user.id:
        return _toast_error("Vous ne pouvez pas émarger votre propre cours.")

    try:
        d = _dt.strptime(request.POST.get('date', ''), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return HttpResponse(status=400)

    session = request.POST.get('session', 'morning')
    if session not in {c[0] for c in SessionType.choices}:
        session = 'morning'

    status = request.POST.get('status', '')
    # Statut vide → dé-marquage : on supprime l'émargement éventuel.
    if status == '':
        TeacherAttendance.objects.filter(
            class_subject=cs, date=d, session=session,
        ).delete()
        return HttpResponse(status=204)
    if status not in {c[0] for c in TeacherAttendanceStatus.choices}:
        return HttpResponse(status=400)

    substitute = None
    if status == 'replaced':
        sub_id = request.POST.get('substitute_id')
        if sub_id:
            substitute = User.objects.filter(
                pk=sub_id, memberships__school=school, memberships__role=UserRole.TEACHER,
            ).first()

    # Heures réelles (« partiel ») : uniquement pour un cours assuré (présent).
    hours = None
    if status == 'present':
        hours = _parse_money(request.POST.get('hours'))  # accepte décimaux via parse
        if hours is not None and hours <= 0:
            hours = None

    TeacherAttendance.objects.update_or_create(
        class_subject=cs, date=d, session=session,
        defaults={
            'teacher': cs.teacher, 'school': school, 'status': status,
            'hours': hours, 'substitute': substitute, 'recorded_by': request.user,
            'note': request.POST.get('note', '').strip()[:200],
        },
    )
    return HttpResponse(status=204)


@login_required
@director_or_emargement_required
def emargement_substitute_search(request):
    """Recherche d'un remplaçant (enseignant de l'école)."""
    from apps.accounts.models import User

    school = get_school(request)
    q  = request.GET.get('q', '').strip()
    cs = request.GET.get('cs', '')
    results = []
    if len(q) >= 2:
        results = list(
            User.objects
            .filter(memberships__school=school, memberships__role=UserRole.TEACHER,
                    full_name__icontains=q, is_active=True)
            .distinct().order_by('full_name')[:8]
        )
    return render(request, 'accounting/partials/substitute_results.html', {'results': results, 'cs': cs})


# ─── Phase 4 — Paie mensuelle ────────────────────────────────────────────────

_MOIS_FR = ['', 'janvier', 'février', 'mars', 'avril', 'mai', 'juin',
            'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']


def _parse_year_month(request):
    from datetime import date as _date
    today = _date.today()
    try:
        year = int(request.GET.get('year') or request.POST.get('year') or today.year)
        month = int(request.GET.get('month') or request.POST.get('month') or today.month)
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month
    return year, month


@login_required
@director_or_accounting_required
def salary_dashboard(request):
    """Preview de la paie d'un mois (permanents + vacataires)."""
    from .services import compute_monthly_salary_preview

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    year, month = _parse_year_month(request)
    data = compute_monthly_salary_preview(school, year, month)

    def _sum(rows):
        total = sum((r['computed_amount'] for r in rows), Decimal('0'))
        paid = sum((r['existing_payment'].amount for r in rows
                    if r['existing_payment'] and r['existing_payment'].status == 'paid'), Decimal('0'))
        unpaid = sum(1 for r in rows if r['status'] == 'unpaid')
        pending = sum(1 for r in rows if r['status'] == 'pending')
        return {'total': total, 'paid': paid, 'restant': total - paid,
                'pct': int(paid / total * 100) if total else 0,
                'unpaid': unpaid, 'pending': pending}

    perm_sum = _sum(data['permanents'])
    vac_sum = _sum(data['vacataires'])

    tab = request.GET.get('tab')
    if tab not in ('perm', 'vac'):
        tab = 'perm' if data['permanents'] else 'vac'

    prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
    next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)

    return render(request, 'accounting/salary_dashboard.html', {
        'school': school, 'year': year, 'month': month, 'tab': tab,
        'month_label': _MOIS_FR[month].capitalize(),
        'permanents': data['permanents'], 'vacataires': data['vacataires'],
        'not_configured': data['not_configured'],
        'perm_sum': perm_sum, 'vac_sum': vac_sum,
        'prev_y': prev_y, 'prev_m': prev_m, 'next_y': next_y, 'next_m': next_m,
    })


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_settings_save(request):
    """Règle la retenue par absence (politique d'école) puis recharge la paie."""
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.urls import reverse

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    val = _parse_money(request.POST.get('absence_deduction'))
    school.absence_deduction = val if val is not None else 0
    school.save(update_fields=['absence_deduction'])
    messages.success(request, 'Retenue par absence mise à jour.')
    return redirect(f"{reverse('accounting:salaires')}?year={year}&month={month}")


def _render_salary_row(request, membership, year, month, toast=None, toast_type='success'):
    from .services import salary_row
    resp = render(request, 'accounting/partials/salary_row.html', {
        'row': salary_row(request.school, membership, year, month),
        'year': year, 'month': month,
    })
    if toast:
        resp['HX-Trigger'] = json.dumps({'showToast': {'message': toast, 'type': toast_type}})
    return resp


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_pay(request):
    """Crée une paie en PENDING (montant recalculé serveur + snapshots). Pré-check double-pay."""
    from django.db import IntegrityError
    from apps.accounts.models import Membership
    from apps.payments.models import PaymentMethod
    from .models import SalaryPayment, EmploymentType
    from .services import compute_vacataire_pay, compute_permanent_deductions

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    try:
        membership_id = int(request.POST.get('membership_id'))
    except (TypeError, ValueError):
        return HttpResponse(status=400)
    m = get_object_or_404(
        Membership.objects.select_related('user', 'employee_profile'),
        id=membership_id, school=school,
    )
    profile = getattr(m, 'employee_profile', None)
    if profile is None or not profile.is_active:
        return _toast_error('Profil de rémunération non configuré.')

    # Pré-check double paiement
    if SalaryPayment.objects.filter(employee=m, year=year, month=month, is_cancelled=False).exists():
        return _toast_error(f'{m.user.full_name} a déjà une paie pour ce mois.')

    # Montant + snapshots recalculés SERVEUR (jamais le client)
    deduction, absence_count = Decimal('0'), 0
    if profile.employment_type == EmploymentType.PERMANENT:
        gross = profile.monthly_salary or Decimal('0')
        d = compute_permanent_deductions(school, year, month).get(m.user_id) or {}
        deduction = d.get('deduction', Decimal('0'))
        absence_count = d.get('absences', 0)
        amount = gross - deduction
        if amount < 0:
            amount = Decimal('0')
        hours, rate = None, None
    else:
        v = compute_vacataire_pay(school, year, month).get(m.user_id) or {}
        hours = v.get('hours', Decimal('0'))
        amount = v.get('amount', Decimal('0'))
        rate = None  # tarif par cours (plus de taux unique)

    method = request.POST.get('payment_method', 'cash')
    if method not in {c[0] for c in PaymentMethod.choices}:
        method = 'cash'

    try:
        SalaryPayment.objects.create(
            employee=m, school=school, year=year, month=month,
            amount=amount, hours=hours, hourly_rate=rate,
            deduction=deduction, absence_count=absence_count,
            status='pending', payment_method=method,
            employee_name=m.user.full_name,
        )
    except IntegrityError:
        return _toast_error('Paiement déjà enregistré pour ce mois.')

    return _render_salary_row(request, m, year, month, toast='Paie créée (en attente de confirmation).')


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_pay_all(request):
    """Initie (PENDING) toutes les paies non encore créées d'un type, montant > 0."""
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.contrib import messages
    from .models import SalaryPayment
    from .services import compute_monthly_salary_preview

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)
    emp_type = request.POST.get('type')
    tab = 'vac' if emp_type == 'vacataire' else 'perm'

    data = compute_monthly_salary_preview(school, year, month)
    rows = data['vacataires'] if emp_type == 'vacataire' else data['permanents']
    created = 0
    for r in rows:
        if r['status'] != 'unpaid' or r['computed_amount'] <= 0:
            continue
        m = r['membership']
        if SalaryPayment.objects.filter(employee=m, year=year, month=month, is_cancelled=False).exists():
            continue
        SalaryPayment.objects.create(
            employee=m, school=school, year=year, month=month,
            amount=r['computed_amount'], hours=r['computed_hours'], hourly_rate=None,
            deduction=r['deduction'], absence_count=r['absences'],
            status='pending', payment_method='cash', employee_name=m.user.full_name,
        )
        created += 1
    messages.success(request, f"{created} paie(s) initiée(s)." if created else "Aucune paie à initier.")
    return redirect(f"{reverse('accounting:salaires')}?year={year}&month={month}&tab={tab}")


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_confirm_all(request):
    """Confirme (PAID) toutes les paies PENDING d'un type."""
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.contrib import messages
    from django.utils import timezone
    from .models import SalaryPayment, EmploymentType

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)
    emp_type = request.POST.get('type')
    tab = 'vac' if emp_type == 'vacataire' else 'perm'
    et = EmploymentType.VACATAIRE if emp_type == 'vacataire' else EmploymentType.PERMANENT

    n = (SalaryPayment.objects
         .filter(school=school, year=year, month=month, status='pending', is_cancelled=False,
                 employee__employee_profile__employment_type=et)
         .update(status='paid', paid_at=timezone.now(), paid_by=request.user))
    messages.success(request, f"{n} paiement(s) confirmé(s)." if n else "Aucune paie à confirmer.")
    return redirect(f"{reverse('accounting:salaires')}?year={year}&month={month}&tab={tab}")


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_confirm(request, payment_id):
    """PENDING → PAID + paid_at/paid_by."""
    from django.utils import timezone
    from .models import SalaryPayment

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    sp = get_object_or_404(
        SalaryPayment.objects.select_related('employee__user', 'employee__employee_profile'),
        pk=payment_id, school=school, is_cancelled=False,
    )
    if sp.status != 'paid':
        sp.status = 'paid'
        sp.paid_at = timezone.now()
        sp.paid_by = request.user
        sp.save(update_fields=['status', 'paid_at', 'paid_by'])
    return _render_salary_row(request, sp.employee, year, month, toast='Paiement confirmé.')


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def salary_cancel(request, payment_id):
    """Annulation soft (is_cancelled=True) → la ligne redevient payable."""
    from .models import SalaryPayment

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        return _toast_error("Veuillez indiquer un motif d'annulation.")

    sp = get_object_or_404(
        SalaryPayment.objects.select_related('employee__user', 'employee__employee_profile'),
        pk=payment_id, school=school, is_cancelled=False,
    )
    sp.is_cancelled = True
    sp.cancelled_at = timezone.now()
    sp.cancelled_by = request.user
    sp.cancellation_reason = reason
    sp.save(update_fields=[
        'is_cancelled', 'cancelled_at', 'cancelled_by', 'cancellation_reason',
    ])
    return _render_salary_row(request, sp.employee, year, month, toast='Paie annulée.', toast_type='info')


@login_required
@director_or_accounting_required
def payslip_pdf(request, payment_id):
    """Fiche de paie PDF (WeasyPrint) — paiements non annulés."""
    from .models import SalaryPayment
    from .services import generate_payslip_pdf

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    sp = get_object_or_404(
        SalaryPayment.objects.select_related('employee__user', 'school'),
        pk=payment_id, school=school, is_cancelled=False,
    )
    pdf = generate_payslip_pdf(sp)
    filename = f'paie_{sp.employee_name.replace(" ", "_")}_{sp.month}_{sp.year}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="{filename}"'
    return resp


# ─── Phase 5 — Dépenses ──────────────────────────────────────────────────────

def _expense_categories(school):
    """Catégories disponibles : globales (school=NULL) + propres à l'école."""
    from django.db.models import Q
    from .models import ExpenseCategory
    return (ExpenseCategory.objects
            .filter(Q(school__isnull=True) | Q(school=school), is_active=True)
            .order_by('-is_default', 'name'))


# Palette déterministe par catégorie (pastille + couleur de barre).
_CAT_PALETTE = [
    ('#FAEEDA', '#BA7517'), ('#E6F1FB', '#185FA5'), ('#EEEDFE', '#534AB7'),
    ('#E1F5EE', '#0F6E56'), ('#FBEAF0', '#993556'), ('#FAECE7', '#993C1D'),
]


def _cat_color(cid):
    return _CAT_PALETTE[(cid or 0) % len(_CAT_PALETTE)]


def _recurring_status(school, year, month):
    """Récurrentes actives + si déjà enregistrées ce mois (occurrence non annulée)."""
    from .models import RecurringExpense, Expense
    recs = list(
        RecurringExpense.objects.filter(school=school, is_active=True).select_related('category')
    )
    reg = set(
        Expense.objects.filter(
            school=school, recurring__in=recs, is_cancelled=False,
            date__year=year, date__month=month,
        ).values_list('recurring_id', flat=True)
    )
    out = []
    for r in recs:
        r.chip_bg, r.chip_fg = _cat_color(r.category_id)
        out.append({'rec': r, 'registered': r.id in reg})
    return out


def _expense_context(request, school, year, month, category_id, method):
    """Liste + totaux + récurrentes d'un mois filtré (mutualisé dashboard + refresh)."""
    from django.db.models import Sum
    from .models import Expense

    # La LISTE inclut les annulées (affichées avec badge) ; les TOTAUX ci-dessous
    # restent sur is_cancelled=False. Actives d'abord, annulées en bas.
    # Groupée par jour : tri par date (actives + annulées d'un même jour ensemble).
    qs = (Expense.objects
          .filter(school=school, date__year=year, date__month=month)
          .select_related('category', 'paid_by', 'cancelled_by')
          .order_by('-date', 'is_cancelled', '-created_at'))
    if category_id:
        qs = qs.filter(category_id=category_id)
    if method:
        qs = qs.filter(payment_method=method)
    expenses = list(qs)
    for e in expenses:
        e.chip_bg, e.chip_fg = _cat_color(e.category_id)

    # Total + répartition par catégorie (sur le mois, hors filtres catégorie/mode)
    base = Expense.objects.filter(school=school, is_cancelled=False, date__year=year, date__month=month)
    total = base.aggregate(s=Sum('amount'))['s'] or 0
    by_cat = list(
        base.values('category__id', 'category__name', 'category__icon')
        .annotate(s=Sum('amount')).order_by('-s')
    )
    for c in by_cat:
        c['pct'] = int(c['s'] / total * 100) if total else 0
        c['chip'], c['color'] = _cat_color(c['category__id'])

    # Delta vs mois précédent.
    py, pm = (year, month - 1) if month > 1 else (year - 1, 12)
    prev_total = (Expense.objects
                  .filter(school=school, is_cancelled=False, date__year=py, date__month=pm)
                  .aggregate(s=Sum('amount'))['s'] or 0)
    delta_pct = None
    if prev_total:
        delta_pct = round((float(total) - float(prev_total)) / float(prev_total) * 100)

    recurrings = _recurring_status(school, year, month)
    return {
        'expenses': expenses, 'total': total, 'by_cat': by_cat,
        'recurrings': recurrings,
        'rec_total': len(recurrings),
        'rec_done': sum(1 for i in recurrings if i['registered']),
        'delta_pct': delta_pct,
        'year': year, 'month': month, 'category_id': category_id, 'method': method,
    }


@login_required
@director_or_accounting_required
def expense_dashboard(request):
    from apps.payments.models import PaymentMethod

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    year, month = _parse_year_month(request)
    category_id = request.GET.get('category') or ''
    method = request.GET.get('method') or ''
    try:
        category_id = int(category_id) if category_id else ''
    except ValueError:
        category_id = ''

    ctx = _expense_context(request, school, year, month, category_id, method)
    prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
    next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)
    ctx.update({
        'school': school, 'categories': _expense_categories(school),
        'methods': PaymentMethod.choices,
        'month_label': _MOIS_FR[month].capitalize(),
        'prev_y': prev_y, 'prev_m': prev_m, 'next_y': next_y, 'next_m': next_m,
    })
    return render(request, 'accounting/expense_dashboard.html', ctx)


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def expense_create(request):
    from datetime import datetime as _dt, date as _date
    from apps.payments.models import PaymentMethod
    from .models import Expense, ExpenseCategory

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    amount = _parse_money(request.POST.get('amount'))
    if amount is None or amount <= 0:
        return _toast_error('Montant invalide.')

    # Catégorie : valider AVANT le filtre (pk='' ou non-numérique → ValueError → 500).
    category_id = request.POST.get('category', '').strip()
    if not category_id.isdigit():
        return _toast_error('Veuillez sélectionner une catégorie.')

    # Résolution catégorie : globale (school=NULL) ou propre à l'école
    from django.db.models import Q
    cat = ExpenseCategory.objects.filter(
        Q(school__isnull=True) | Q(school=school),
        pk=category_id, is_active=True,
    ).first()
    if cat is None:
        return _toast_error('Catégorie invalide.')

    try:
        d = _dt.strptime(request.POST.get('date', ''), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        d = _date.today()

    method = request.POST.get('payment_method', 'cash')
    if method not in {c[0] for c in PaymentMethod.choices}:
        method = 'cash'

    Expense.objects.create(
        school=school, category=cat, amount=amount, date=d,
        description=request.POST.get('description', '').strip()[:300],
        payment_method=method, paid_by=request.user,
    )

    # Re-render liste + totaux du mois affiché
    year, month = _parse_year_month(request)
    ctx = _expense_context(request, school, year, month, '', '')
    resp = render(request, 'accounting/partials/expense_list.html', ctx)
    resp['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Dépense enregistrée.', 'type': 'success'},
        'close-expense-panel': True,
    })
    return resp


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def expense_cancel(request, expense_id):
    from .models import Expense

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        return _toast_error("Veuillez indiquer un motif d'annulation.")

    exp = get_object_or_404(Expense, pk=expense_id, school=school, is_cancelled=False)
    exp.is_cancelled = True
    exp.cancelled_at = timezone.now()
    exp.cancelled_by = request.user
    exp.cancellation_reason = reason
    exp.save(update_fields=[
        'is_cancelled', 'cancelled_at', 'cancelled_by', 'cancellation_reason',
    ])

    year, month = _parse_year_month(request)
    ctx = _expense_context(request, school, year, month, '', '')
    resp = render(request, 'accounting/partials/expense_list.html', ctx)
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Dépense annulée.', 'type': 'info'}})
    return resp


def _expense_list_response(request, school, year, month, message, extra_trigger=None):
    """Re-render #expense-content + toast (mutualisé récurrentes)."""
    ctx = _expense_context(request, school, year, month, '', '')
    resp = render(request, 'accounting/partials/expense_list.html', ctx)
    trig = {'showToast': {'message': message, 'type': 'success'}}
    if extra_trigger:
        trig.update(extra_trigger)
    resp['HX-Trigger'] = json.dumps(trig)
    return resp


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def recurring_register(request):
    """Enregistre une dépense récurrente pour le mois affiché (1 clic)."""
    from datetime import date as _date
    from .models import RecurringExpense, Expense

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    try:
        rec = RecurringExpense.objects.select_related('category').get(
            pk=int(request.POST.get('recurring_id')), school=school, is_active=True,
        )
    except (TypeError, ValueError, RecurringExpense.DoesNotExist):
        return _toast_error('Récurrente introuvable.')

    already = Expense.objects.filter(
        school=school, recurring=rec, is_cancelled=False,
        date__year=year, date__month=month,
    ).exists()
    if not already:
        today = _date.today()
        d = today if (year, month) == (today.year, today.month) else _date(year, month, 1)
        Expense.objects.create(
            school=school, category=rec.category, recurring=rec,
            amount=rec.amount, date=d, description=rec.label,
            payment_method=rec.payment_method, paid_by=request.user,
        )
    return _expense_list_response(request, school, year, month, 'Dépense récurrente enregistrée.')


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def recurring_create(request):
    """Ajoute un modèle de dépense récurrente."""
    from django.db.models import Q
    from .models import RecurringExpense, ExpenseCategory

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    amount = _parse_money(request.POST.get('amount'))
    if amount is None or amount <= 0:
        return _toast_error('Montant invalide.')
    cat_id = (request.POST.get('category') or '').strip()
    if not cat_id.isdigit():
        return _toast_error('Veuillez sélectionner une catégorie.')
    cat = ExpenseCategory.objects.filter(
        Q(school__isnull=True) | Q(school=school), pk=cat_id, is_active=True,
    ).first()
    if cat is None:
        return _toast_error('Catégorie invalide.')

    from apps.payments.models import PaymentMethod
    method = request.POST.get('payment_method', 'cash')
    if method not in {c[0] for c in PaymentMethod.choices}:
        method = 'cash'

    RecurringExpense.objects.create(
        school=school, category=cat, amount=amount,
        label=request.POST.get('label', '').strip()[:200], payment_method=method,
    )
    return _expense_list_response(
        request, school, year, month, 'Récurrente ajoutée.',
        extra_trigger={'close-recurring-panel': True},
    )


@login_required
@director_or_accounting_required
@require_http_methods(['POST'])
def recurring_delete(request, rec_id):
    """Désactive une récurrente (l'historique des occurrences est conservé)."""
    from .models import RecurringExpense

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)
    year, month = _parse_year_month(request)

    RecurringExpense.objects.filter(pk=rec_id, school=school).update(is_active=False)
    return _expense_list_response(request, school, year, month, 'Récurrente supprimée.')


# ─── Phase 6 — Bilan (fusionné dans la Vue d'ensemble) ───────────────────────

@login_required
@director_or_accounting_required
def bilan_dashboard(request):
    """Fusionné dans « Finances · Vue d'ensemble » → redirection (compat liens)."""
    from django.shortcuts import redirect
    from django.urls import reverse
    year, month = _parse_year_month(request)
    return redirect(f"{reverse('accounting:dashboard')}?year={year}&month={month}")


@login_required
@director_or_accounting_required
def bilan_export_excel(request):
    """Export Excel du bilan d'un mois (openpyxl)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from .services import compute_monthly_balance

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    year, month = _parse_year_month(request)
    b = compute_monthly_balance(school, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Trésorerie {month}-{year}'[:31]

    bold = Font(bold=True)
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

    ws.append([f'{school.name} — Trésorerie {_MOIS_FR[month].capitalize()} {year}'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(['Poste', 'Montant (FCFA)'])
    for cell in ws[ws.max_row]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal='center')
    ws.append(['Revenus (paiements élèves)', float(b['revenus'])])
    ws.append(['Salaires payés', float(b['salaires'])])
    ws.append(['Dépenses', float(b['depenses'])])
    ws.append(['Total charges', float(b['charges'])])
    r = ws.max_row + 1
    ws.append(['Résultat net', float(b['resultat'])])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = PatternFill(
            start_color='DCFCE7' if b['resultat'] >= 0 else 'FEE2E2',
            end_color='DCFCE7' if b['resultat'] >= 0 else 'FEE2E2', fill_type='solid',
        )

    ws.append([])
    ws.append(['Détail des dépenses par catégorie'])
    ws[ws.max_row][0].font = bold
    ws.append(['Catégorie', 'Montant (FCFA)'])
    for cell in ws[ws.max_row]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal='center')
    for c in b['by_cat']:
        ws.append([c['category__name'], float(c['total'])])

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18

    filename = f'tresorerie_{school.name.replace(" ", "_")}_{month}_{year}.xlsx'
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# ─── Phase 7 — Dashboard comptabilité ────────────────────────────────────────

@login_required
@director_or_accounting_required
def accounting_dashboard(request):
    """Finances · Vue d'ensemble — fusion résultat net + bilan : KPIs, graphe 6 mois,
    dépenses par catégorie, alertes, raccourcis, export. Réutilise les calculs."""
    from datetime import date as _date
    from .models import SalaryPayment, SalaryStatus, Expense, TeacherAttendance
    from .services import compute_monthly_balance, compute_balance_series
    from apps.schools.models import ClassSubject

    school = get_school(request)
    if not school.accounting_enabled:
        return HttpResponse(status=403)

    today = _date.today()
    year, month = _parse_year_month(request)
    is_current = (year, month) == (today.year, today.month)

    balance = compute_monthly_balance(school, year, month)
    series  = compute_balance_series(school, year, month, n=6)

    salaires_en_attente = SalaryPayment.objects.filter(
        school=school, status=SalaryStatus.PENDING, is_cancelled=False,
    ).count()

    # Alerte émargement : seulement quand on regarde le mois courant.
    non_emarges = 0
    if is_current:
        total_cours = ClassSubject.objects.filter(
            school_class__school=school, school_class__is_active=True,
            is_active=True, teacher__isnull=False,
        ).count()
        emarges = TeacherAttendance.objects.filter(school=school, date=today).count()
        non_emarges = max(0, total_cours - emarges)

    # Barre revenus/charges (proportion).
    rev = float(balance['revenus'] or 0)
    chg = float(balance['charges'] or 0)
    charges_pct = min(100, round(chg / rev * 100)) if rev > 0 else (100 if chg > 0 else 0)
    kept_pct = max(0, 100 - charges_pct)

    prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
    next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)

    chart_labels   = [f"{_MOIS_FR[s['month']][:4].capitalize()}" for s in series]
    chart_revenus  = [s['revenus']  for s in series]
    chart_charges  = [s['charges']  for s in series]
    chart_resultat = [s['resultat'] for s in series]

    return render(request, 'accounting/dashboard.html', {
        'school': school, 'today': today, 'is_current': is_current,
        'month_label': _MOIS_FR[month].capitalize(), 'year': year, 'month': month,
        'prev_y': prev_y, 'prev_m': prev_m, 'next_y': next_y, 'next_m': next_m,
        'b': balance, 'charges_pct': charges_pct, 'kept_pct': kept_pct,
        'salaires_en_attente': salaires_en_attente,
        'non_emarges': non_emarges,
        'chart_labels': chart_labels, 'chart_revenus': chart_revenus,
        'chart_charges': chart_charges, 'chart_resultat': chart_resultat,
    })
