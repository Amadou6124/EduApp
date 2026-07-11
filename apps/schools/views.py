import csv
import io
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.urls import reverse

from apps.accounts.models import UserRole
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods

from .models import SchoolClass, School, EducationLevel, SchoolAnnouncement
from .forms import SchoolClassForm
from .cockpit import build_class_cockpit
from apps.core.mixins import get_school, director_or_staff_required
from apps.core.text import norm_name

# Correspondance libellés Excel → valeurs modèle
LEVEL_LABELS = {
    'prescolaire':                EducationLevel.PRESCOLAIRE,
    'préscolaire':                EducationLevel.PRESCOLAIRE,
    'fondamental1':               EducationLevel.FONDAMENTAL_1,
    'fondamental_1':              EducationLevel.FONDAMENTAL_1,
    'fondamental 1er cycle':      EducationLevel.FONDAMENTAL_1,
    'primaire':                   EducationLevel.FONDAMENTAL_1,
    'fondamental2':               EducationLevel.FONDAMENTAL_2,
    'fondamental_2':              EducationLevel.FONDAMENTAL_2,
    'fondamental 2ème cycle':     EducationLevel.FONDAMENTAL_2,
    'fondamental 2eme cycle':     EducationLevel.FONDAMENTAL_2,
    'college':                    EducationLevel.FONDAMENTAL_2,
    'collège':                    EducationLevel.FONDAMENTAL_2,
    'secondaire':                 EducationLevel.SECONDAIRE_GEN,
    'secondaire_gen':             EducationLevel.SECONDAIRE_GEN,
    'secondaire général':         EducationLevel.SECONDAIRE_GEN,
    'secondaire general':         EducationLevel.SECONDAIRE_GEN,
    'lycee':                      EducationLevel.SECONDAIRE_GEN,
    'lycée':                      EducationLevel.SECONDAIRE_GEN,
    'secondaire_pro':             EducationLevel.SECONDAIRE_PRO,
    'secondaire professionnel':   EducationLevel.SECONDAIRE_PRO,
    'cap':                        EducationLevel.SECONDAIRE_PRO,
    'universite':                 EducationLevel.SUPERIEUR,
    'université':                 EducationLevel.SUPERIEUR,
    'superieur':                  EducationLevel.SUPERIEUR,
    'supérieur':                  EducationLevel.SUPERIEUR,
    'enseignement supérieur':     EducationLevel.SUPERIEUR,
    'enseignement superieur':     EducationLevel.SUPERIEUR,
}
LEVEL_DISPLAY = {
    EducationLevel.PRESCOLAIRE:    'Préscolaire',
    EducationLevel.FONDAMENTAL_1:  'Fondamental 1er Cycle',
    EducationLevel.FONDAMENTAL_2:  'Fondamental 2ème Cycle',
    EducationLevel.SECONDAIRE_GEN: 'Secondaire Général',
    EducationLevel.SECONDAIRE_PRO: 'Secondaire Professionnel',
    EducationLevel.SUPERIEUR:      'Enseignement Supérieur',
}


def _classes_qs(school):
    return (
        SchoolClass.objects
        .filter(school=school, is_active=True)
        .select_related('school')
        .annotate(
            student_count=Count('students', filter=Q(students__is_active=True)),
            boys=Count('students',  filter=Q(students__is_active=True, students__gender='M')),
            girls=Count('students', filter=Q(students__is_active=True, students__gender='F')),
        )
    )


def compute_class_stats(classes):
    """Agrégats liste classes : effectif, parité, et métrique de capacité honnête
    (places restantes + classes complètes — pas une moyenne plafonnée trompeuse)."""
    with_cap = [c for c in classes if c.max_capacity]
    return {
        'total_classes':     len(classes),
        'total_students':    sum(c.student_count for c in classes),
        'boys':              sum(c.boys for c in classes),
        'girls':             sum(c.girls for c in classes),
        'places_restantes':  sum(max(c.max_capacity - c.student_count, 0) for c in with_cap),
        'classes_completes': sum(1 for c in with_cap if c.student_count >= c.max_capacity),
    }


def _filter_classes(request, school):
    """Classes filtrées : recherche (nom) × niveau × disponibilité (places libres)."""
    qs = _classes_qs(school)
    q     = request.GET.get('q', '').strip()
    level = request.GET.get('level', '')
    if q:
        qs = qs.filter(name__icontains=q)
    if level:
        qs = qs.filter(level=level)
    classes = sorted(qs, key=lambda c: (c.level, c.name))
    if request.GET.get('dispo') in ('1', 'true', 'on'):
        classes = [c for c in classes if not c.is_full]
    return classes


@login_required
def class_list(request):
    if request.user.role == UserRole.TEACHER:
        return redirect('teacher:dashboard')
    school = get_school(request)
    all_classes = list(_classes_qs(school))
    return render(request, 'schools/class_list.html', {
        'classes':        _filter_classes(request, school),
        'form':           SchoolClassForm(),
        'school':         school,
        'cstats':         compute_class_stats(all_classes),
        'levels':         EducationLevel.choices,
        'q':              request.GET.get('q', ''),
        'level':          request.GET.get('level', ''),
        'dispo':          request.GET.get('dispo', ''),
    })


@login_required
def class_detail(request, class_id):
    """Cockpit de classe : KPIs, élèves à risque, moyenne par matière, roster, matières/profs."""
    if request.user.role == UserRole.TEACHER:
        return redirect('teacher:dashboard')
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)
    return render(request, 'schools/class_detail.html', {
        'school_class': school_class,
        'cockpit':      build_class_cockpit(school, school_class),
        'page_title':   school_class.name,
        **_timetable_ctx(school, school_class),   # onglet « Emploi du temps »
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def class_create(request):
    school = get_school(request)

    form = SchoolClassForm(request.POST)
    if not form.is_valid():
        if request.htmx:
            return render(request, 'schools/partials/class_form_fields.html', {'form': form})
        return render(request, 'schools/class_list.html', {
            'form': form, 'school': school, 'classes': list(_classes_qs(school)),
        })

    cd = form.cleaned_data

    # Réactivation d'une classe soft-deletée de même nom (insensible casse ET
    # accents via normalisation Python), sinon création. cleaned_data → max_capacity
    # vide = None (pas de ValueError).
    target = norm_name(cd['name'])
    existing = next(
        (sc for sc in SchoolClass.objects.filter(school=school, is_active=False)
         if norm_name(sc.name) == target),
        None,
    )
    try:
        with transaction.atomic():
            if existing:
                existing.is_active    = True
                existing.name         = cd['name']      # rafraîchit la casse saisie
                existing.level        = cd['level']
                existing.annual_fee   = cd['annual_fee']
                existing.max_capacity = cd['max_capacity']
                existing.save()
                message = _('Classe réactivée avec succès.')
            else:
                school_class = form.save(commit=False)
                school_class.school = school
                school_class.save()
                message = _('Classe créée avec succès.')
    except IntegrityError:
        # Contrainte partielle unique_active_class_per_school (non validée par le
        # ModelForm) : un doublon ACTIF du même nom existe déjà.
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({'showToast': {
            'message': "Une classe avec ce nom existe déjà dans cet établissement.",
            'type': 'error',
        }})
        return resp

    # Réponse succès (HTMX : liste rafraîchie + toast ; sinon page complète).
    if request.htmx:
        classes = list(_classes_qs(school))
        resp = render(request, 'schools/partials/class_list_refresh.html', {
            'classes': classes,
            'form': SchoolClassForm(),
            'success_message': message,
            'cstats': compute_class_stats(classes),
        })
        resp['HX-Trigger'] = json.dumps({'close-add-modal': True, 'showToast': {'message': str(message), 'type': 'success'}})
        return resp

    return render(request, 'schools/class_list.html', {
        'form': SchoolClassForm(), 'school': school, 'classes': list(_classes_qs(school)),
    })


@login_required
def class_edit_form(request, class_id):
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    form = SchoolClassForm(instance=school_class)
    return render(request, 'schools/partials/class_edit_row.html', {
        'form': form,
        'school_class': school_class,
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def class_update(request, class_id):
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    form = SchoolClassForm(request.POST, instance=school_class)

    if form.is_valid():
        try:
            with transaction.atomic():
                form.save()
        except IntegrityError:
            # Renommage vers un nom déjà porté par une classe ACTIVE (contrainte
            # partielle unique_active_class_per_school, non vue par le ModelForm).
            form.add_error('name', "Une classe active porte déjà ce nom.")
        else:
            # Le modal cible #modal-edit-content (toujours présent) ; la ligne du
            # tableau se met à jour en OOB (mise à jour sur la LISTE, inoffensif sur
            # la page DÉTAIL où elle n'existe pas). close-edit-modal ferme le modal
            # (et recharge la page détail) → la modif est TOUJOURS visible.
            resp = render(request, 'schools/partials/class_row.html', {
                'school_class': school_class,
                'success': True,
                'oob': True,
            })
            resp['HX-Trigger'] = json.dumps({
                'close-edit-modal': True,
                'showToast': {'message': f'Classe {school_class.name} modifiée.', 'type': 'success'},
            })
            return resp

    # Erreur (form invalide OU doublon) → réafficher le MODAL avec les erreurs
    # (swap dans #modal-edit-content, la cible du formulaire).
    return render(request, 'schools/partials/class_edit_modal.html', {
        'form': form,
        'school_class': school_class,
    })


@login_required
def class_search(request):
    school = get_school(request)
    return render(request, 'schools/partials/class_table_body.html', {
        'classes': _filter_classes(request, school),
        'q':     request.GET.get('q', ''),
        'level': request.GET.get('level', ''),
        'dispo': request.GET.get('dispo', ''),
    })


@login_required
def class_edit_modal(request, class_id):
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    form = SchoolClassForm(instance=school_class)
    return render(request, 'schools/partials/class_edit_modal.html', {
        'form': form,
        'school_class': school_class,
    })


@login_required
def class_row(request, class_id):
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    return render(request, 'schools/partials/class_row.html', {
        'school_class': school_class,
    })


@login_required
@director_or_staff_required
def class_import_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Classes'

    # En-têtes (* = obligatoire). Le parser lit par position, le renommage est sûr.
    headers = ['Nom de la classe *', 'Niveau *', 'Frais annuels FCFA *', 'Capacité max']
    ws.append(headers)
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 26

    # Lignes d'exemple réalistes (système éducatif malien). À remplacer par vos classes.
    ws.append(['1ère année A',   'Fondamental 1er Cycle',  75000,  50])
    ws.append(['6ème année',     'Fondamental 1er Cycle',  90000,  45])
    ws.append(['9ème année B',   'Fondamental 2ème Cycle', 120000, 40])
    ws.append(['11ème Sciences', 'Secondaire Général',     150000, 35])
    example_font = Font(italic=True, color='9CA3AF')
    for row in ws.iter_rows(min_row=2, max_row=5):
        for cell in row:
            cell.font = example_font

    # Feuille « Instructions » séparée (n'interfère pas avec l'import, qui ne lit
    # que la feuille active « Classes »).
    helpws = wb.create_sheet('Instructions')
    helpws.column_dimensions['A'].width = 40
    helpws.append(['Comment remplir ce modèle'])
    helpws['A1'].font = Font(bold=True, size=12, color='1E3A5F')
    helpws.append([''])
    helpws.append(['Colonnes obligatoires (*) : Nom de la classe, Niveau, Frais annuels FCFA'])
    helpws.append(['Colonne optionnelle : Capacité max (laisser vide si non gérée)'])
    helpws.append(['Frais annuels : montant en FCFA, chiffres uniquement (ex : 90000)'])
    helpws.append([''])
    helpws.append(['Valeurs acceptées pour la colonne « Niveau » :'])
    helpws['A7'].font = Font(bold=True, color='1E3A5F')
    for label in [
        'Préscolaire',
        'Fondamental 1er Cycle',
        'Fondamental 2ème Cycle',
        'Secondaire Général',
        'Secondaire Professionnel',
        'Enseignement Supérieur',
    ]:
        helpws.append([label])
    helpws.append([''])
    helpws.append(['Astuce : supprimez les 4 lignes d\'exemple avant d\'importer vos données.'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_classes.xlsx"'
    wb.save(response)
    return response


def _parse_import_rows(file_obj, filename):
    rows = []
    errors = []

    try:
        if filename.lower().endswith('.csv'):
            content = file_obj.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            next(reader, None)
            raw_rows = list(reader)
        else:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            raw_rows = [
                [str(cell.value).strip() if cell.value is not None else '' for cell in row]
                for row in ws.iter_rows(min_row=2)
            ]
    except Exception as exc:
        errors.append(f'Impossible de lire le fichier : {exc}')
        return rows, errors

    for line_num, raw in enumerate(raw_rows, start=2):
        if not any(raw):
            continue
        # Colonnes : Nom | Niveau | Frais | Capacité
        cols = (raw + ['', '', '', ''])[:4]
        name, level_raw, fee_raw, capacity_raw = [c.strip() for c in cols]

        row_errors = []

        if not name:
            row_errors.append('Nom manquant')

        level_key = level_raw.lower().replace('\xa0', ' ').strip()
        level = LEVEL_LABELS.get(level_key)
        if not level:
            row_errors.append(f'Niveau "{level_raw}" non reconnu')

        try:
            annual_fee = int(float(fee_raw.replace(' ', '').replace('\xa0', '')))
            if annual_fee < 0:
                raise ValueError
        except (ValueError, AttributeError):
            annual_fee = None
            row_errors.append('Frais annuels invalides')

        try:
            max_capacity = int(float(capacity_raw)) if capacity_raw else None
        except ValueError:
            max_capacity = None
            row_errors.append('Capacité invalide')

        if row_errors:
            errors.append({'line': line_num, 'name': name or '—', 'errors': row_errors})
        else:
            rows.append({
                'name': name,
                'level': level,
                'level_display': LEVEL_DISPLAY[level],
                'annual_fee': annual_fee,
                'max_capacity': max_capacity,
            })

    return rows, errors


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def class_import_preview(request):
    file_obj = request.FILES.get('import_file')
    if not file_obj:
        return HttpResponse('<p class="text-red-600 text-sm">Aucun fichier sélectionné.</p>')

    rows, errors = _parse_import_rows(file_obj, file_obj.name)
    school = get_school(request)

    # Détecter les doublons avec les classes existantes
    existing_names = set(
        SchoolClass.objects.filter(school=school, is_active=True).values_list('name', flat=True)
    )
    duplicates = [r['name'] for r in rows if r['name'] in existing_names]

    return render(request, 'schools/partials/class_import_preview.html', {
        'rows': rows,
        'parse_errors': errors,
        'duplicates': duplicates,
        'rows_json': json.dumps(rows),
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def class_import_confirm(request):
    school = get_school(request)
    rows_json = request.POST.get('rows_data', '[]')

    try:
        rows = json.loads(rows_json)
    except json.JSONDecodeError:
        return HttpResponse('<p class="text-red-600 text-sm">Données invalides.</p>')

    created, skipped, reactivated = 0, 0, 0
    for row in rows:
        obj, was_created = SchoolClass.objects.get_or_create(
            school=school,
            name=row['name'],
            defaults={
                'level': row['level'],
                'annual_fee': row['annual_fee'],
                'max_capacity': row.get('max_capacity'),
                'is_active': True,
            },
        )
        if was_created:
            created += 1
        elif not obj.is_active:
            obj.is_active = True
            obj.level = row['level']
            obj.annual_fee = row['annual_fee']
            obj.max_capacity = row.get('max_capacity')
            obj.save()
            reactivated += 1
        else:
            skipped += 1

    msg = f'{created} classe(s) importée(s).'
    if reactivated:
        msg += f' {reactivated} réactivée(s).'
    if skipped:
        msg += f' {skipped} ignorée(s) (doublon).'

    classes = list(_classes_qs(school))

    resp = render(request, 'schools/partials/class_list_refresh.html', {
        'classes': classes,
        'success_message': msg,
        'cstats': compute_class_stats(classes),
    })
    resp['HX-Trigger'] = json.dumps({'close-import-modal': True, 'showToast': {'message': msg, 'type': 'success'}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['DELETE'])
def class_delete(request, class_id):
    school = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    student_count = school_class.get_student_count()
    if student_count > 0:
        response = HttpResponse(status=422)
        response['HX-Trigger'] = json.dumps({
            'showToast': {
                'message': f'Impossible : {student_count} élève(s) inscrit(s) dans cette classe.',
                'type': 'error',
            }
        })
        return response
    name = school_class.name
    school_class.is_active = False
    school_class.save()
    response = HttpResponse('')
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': f'Classe {name} supprimée.', 'type': 'success'}
    })
    return response


# ─────────────────────────────────────────────────────────────────────
# ANNONCES
# ─────────────────────────────────────────────────────────────────────

def _annotate_read_receipts(announcements):
    """Annote chaque annonce publiée : reach (parents notifiés) · read_count · read_pct.
    Les notifs sont liées à l'annonce (target=ann) → accusés de lecture fiables."""
    from django.contrib.contenttypes.models import ContentType
    from apps.notifications.models import Notification

    ids = [a.pk for a in announcements if a.is_published]
    reach, read = {}, {}
    if ids:
        ct = ContentType.objects.get_for_model(SchoolAnnouncement)
        reach = dict(
            Notification.objects.filter(content_type=ct, object_id__in=ids)
            .values('object_id').annotate(n=Count('id')).values_list('object_id', 'n')
        )
        read = dict(
            Notification.objects.filter(content_type=ct, object_id__in=ids, is_read=True)
            .values('object_id').annotate(n=Count('id')).values_list('object_id', 'n')
        )
    for a in announcements:
        a.reach = reach.get(a.pk, 0)
        a.read_count = read.get(a.pk, 0)
        a.read_pct = round(a.read_count / a.reach * 100) if a.reach else 0


def _publish_announcement(ann):
    """Publie une annonce + notifie les parents ciblés (target=ann pour les accusés de lecture)."""
    from django.utils import timezone
    from apps.notifications.services import notify_bulk, notify_guardians
    from apps.notifications.models import NotificationCategory
    from apps.students.models import StudentGuardian

    if ann.is_published:
        return
    ann.is_published = True
    ann.published_at = timezone.now()
    ann.save(update_fields=['is_published', 'published_at'])

    school     = ann.school
    notif_url  = reverse('parent:annonces')
    notif_body = ann.body[:150]

    if ann.audience == 'school':
        ids = list(StudentGuardian.objects.filter(student__school=school).values_list('guardian_id', flat=True).distinct())
        notify_bulk(ids, school, NotificationCategory.INFO, ann.title, notif_body, notif_url, target=ann)
    elif ann.audience == 'class' and ann.target_class:
        ids = list(StudentGuardian.objects.filter(student__school_class=ann.target_class).values_list('guardian_id', flat=True).distinct())
        notify_bulk(ids, school, NotificationCategory.INFO, ann.title, notif_body, notif_url, target=ann)
    elif ann.audience == 'student' and ann.target_student:
        notify_guardians(ann.target_student, NotificationCategory.INFO, ann.title, notif_body, notif_url, target=ann)


@login_required
@director_or_staff_required
def announcement_list(request):
    school = get_school(request)
    from apps.students.models import Student

    qs = (
        SchoolAnnouncement.objects
        .filter(school=school)
        .select_related('author', 'target_class', 'target_student')
        .order_by('-created_at')
    )
    status = request.GET.get('status', 'all')
    if status == 'published':
        qs = qs.filter(is_published=True)
    elif status == 'draft':
        qs = qs.filter(is_published=False)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(body__icontains=q))

    announcements = list(qs)
    _annotate_read_receipts(announcements)

    ctx = {
        'announcements': announcements,
        'classes':       SchoolClass.objects.filter(school=school, is_active=True).order_by('name'),
        'students':      Student.objects.filter(school=school, is_active=True).select_related('school_class').order_by('full_name'),
        'status':        status,
        'q':             q,
    }
    if request.htmx:
        return render(request, 'schools/announcements/partials/list_body.html', ctx)
    return render(request, 'schools/announcements/list.html', ctx)


def _err_toast(msg):
    resp = HttpResponse(status=422)
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': msg, 'type': 'error'}})
    return resp


def _parse_announcement_form(request, school):
    """Valide le formulaire annonce. Renvoie (data, None) ou (None, error_response)."""
    from apps.students.models import Student
    title    = request.POST.get('title', '').strip()
    body     = request.POST.get('body', '').strip()
    audience = request.POST.get('audience', 'school')
    if not title:
        return None, _err_toast('Le titre est requis.')
    if not body:
        return None, _err_toast('Le contenu est requis.')
    if audience not in ('school', 'class', 'student'):
        audience = 'school'
    target_class = target_student = None
    if audience == 'class':
        class_id = request.POST.get('target_class_id', '').strip()
        if not class_id:
            return None, _err_toast('Sélectionnez une classe.')
        target_class = get_object_or_404(SchoolClass, id=class_id, school=school)
    if audience == 'student':
        student_id = request.POST.get('target_student_id', '').strip()
        if not student_id:
            return None, _err_toast('Sélectionnez un élève.')
        target_student = get_object_or_404(Student, id=student_id, school=school)
    return {
        'title': title, 'body': body, 'audience': audience,
        'target_class': target_class, 'target_student': target_student,
    }, None


@login_required
def announcement_form(request):
    """Formulaire de composition (nouveau ou édition de brouillon) — chargé dans le panneau."""
    school = get_school(request)
    from apps.students.models import Student
    ann = None
    aid = request.GET.get('id')
    if aid:
        ann = get_object_or_404(SchoolAnnouncement, pk=aid, school=school, is_published=False)
    return render(request, 'schools/announcements/partials/announcement_form.html', {
        'ann':      ann,
        'classes':  SchoolClass.objects.filter(school=school, is_active=True).order_by('name'),
        'students': Student.objects.filter(school=school, is_active=True).select_related('school_class').order_by('full_name'),
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def announcement_create(request):
    school = get_school(request)
    data, err = _parse_announcement_form(request, school)
    if err:
        return err

    ann = SchoolAnnouncement.objects.create(school=school, author=request.user, **data)
    publish_now = request.POST.get('publish_now') in ('1', 'true', 'on')
    if publish_now:
        _publish_announcement(ann)
    _annotate_read_receipts([ann])

    resp = render(request, 'schools/announcements/partials/announcement_card.html', {'ann': ann})
    resp['HX-Trigger'] = json.dumps({
        'close-announcement-panel': True,
        'showToast': {'message': ('Annonce publiée et envoyée.' if publish_now else 'Brouillon créé.'), 'type': 'success'},
    })
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def announcement_update(request, pk):
    school = get_school(request)
    ann    = get_object_or_404(SchoolAnnouncement, pk=pk, school=school, is_published=False)
    data, err = _parse_announcement_form(request, school)
    if err:
        return err

    for k, v in data.items():
        setattr(ann, k, v)
    ann.save()

    publish_now = request.POST.get('publish_now') in ('1', 'true', 'on')
    if publish_now:
        _publish_announcement(ann)
    _annotate_read_receipts([ann])

    resp = render(request, 'schools/announcements/partials/announcement_card.html', {'ann': ann})
    resp['HX-Trigger'] = json.dumps({
        'close-announcement-panel': True,
        'showToast': {'message': ('Brouillon publié.' if publish_now else 'Brouillon mis à jour.'), 'type': 'success'},
    })
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def announcement_publish(request, pk):
    school = get_school(request)
    ann    = get_object_or_404(SchoolAnnouncement, pk=pk, school=school)

    if ann.is_published:
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Annonce déjà publiée.', 'type': 'error'}})
        return resp

    _publish_announcement(ann)
    _annotate_read_receipts([ann])

    resp = render(request, 'schools/announcements/partials/announcement_card.html', {'ann': ann})
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Annonce publiée et envoyée.', 'type': 'success'}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def announcement_delete(request, pk):
    school = get_school(request)
    ann    = get_object_or_404(SchoolAnnouncement, pk=pk, school=school)

    if ann.is_published:
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Impossible de supprimer une annonce publiée.', 'type': 'error'}})
        return resp

    ann.delete()
    resp = HttpResponse('')
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Brouillon supprimé.', 'type': 'success'}})
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# EMPLOI DU TEMPS — grille par classe (créneaux + pauses)
# ══════════════════════════════════════════════════════════════════════════════
# La grille dessine les blocs en positionnement PROPORTIONNEL (1 px = 1 minute) :
# les heures non rondes (11h15, 13h30…) tombent naturellement au bon endroit.

_GRID_START_MIN = 7 * 60 + 30   # 7h30
_GRID_END_MIN   = 18 * 60       # 18h00


def _min_of(t):
    return t.hour * 60 + t.minute


def _timetable_ctx(school, school_class):
    """Contexte de la grille : créneaux positionnés par jour, pauses, cours de la
    classe (pour le formulaire), repères d'heures. Année active uniquement."""
    from .models import CourseSlot, SchoolBreak, ClassSubject, Weekday
    from .periods import active_year_for

    year = active_year_for(school)
    days = list(Weekday.choices)

    slots_by_day = {d: [] for d, _label in days}
    scheduled_cs_ids = set()   # cours ayant AU MOINS un créneau (→ indicateur d'état)
    if year:
        qs = (CourseSlot.objects
              .filter(school_year=year, class_subject__school_class=school_class)
              .select_related('class_subject__subject', 'class_subject__teacher')
              .order_by('day', 'start_time'))
        for s in qs:
            scheduled_cs_ids.add(s.class_subject_id)
            top    = max(_min_of(s.start_time) - _GRID_START_MIN, 0)
            height = max(_min_of(s.end_time) - max(_min_of(s.start_time), _GRID_START_MIN), 18)
            slots_by_day.setdefault(s.day, []).append({
                'slot': s, 'top': top, 'height': height,
                'subject': s.class_subject.subject,
                'teacher': s.class_subject.teacher,
            })

    breaks = list(SchoolBreak.objects.filter(school=school))
    # Dimanche visible SEULEMENT sur contenu explicitement dominical (créneau ou pause
    # « Dimanche ») — une pause « tous les jours » ne doit pas faire surgir la colonne.
    sunday_visible = bool(slots_by_day.get(6)) or any(b.day == 6 for b in breaks)
    visible_days = [d for d, _l in days if d != 6 or sunday_visible]

    breaks_by_day = {d: [] for d, _label in days}
    for b in breaks:
        top    = max(_min_of(b.start_time) - _GRID_START_MIN, 0)
        height = max(_min_of(b.end_time) - max(_min_of(b.start_time), _GRID_START_MIN), 12)
        target_days = [b.day] if b.day is not None else visible_days
        for d in target_days:
            breaks_by_day.setdefault(d, []).append({'brk': b, 'top': top, 'height': height})

    class_subjects = list(
        ClassSubject.objects.filter(school_class=school_class, is_active=True)
        .select_related('subject', 'teacher').order_by('order', 'subject__name')
    )
    # Durée par défaut (minutes) par cours → pré-remplissage de l'heure de fin côté JS.
    durations = {cs.id: int((cs.duration_hours or 2) * 60) for cs in class_subjects}

    hour_marks = []   # repères 8h → 17h (positions en px depuis 7h30)
    for h in range(8, 18):
        hour_marks.append({'label': f'{h}h', 'top': h * 60 - _GRID_START_MIN})

    # Chaque jour porte directement ses créneaux/pauses → itération simple au template
    # (pas de filtre custom pour indexer un dict par variable).
    # Dimanche (6) : colonne affichée SEULEMENT si utilisée (franco-arabes) — mais
    # toujours proposé dans les formulaires (weekday_choices complet).
    day_columns = [
        {'val': d, 'label': label,
         'slots': slots_by_day.get(d, []), 'breaks': breaks_by_day.get(d, [])}
        for d, label in days if d in visible_days
    ]

    # Indicateur d'état LOCAL (uniquement des trous que l'app SAIT détecter — jamais
    # de « il manque EPS » : aucun référentiel de programme n'existe).
    edt_no_teacher = sum(1 for cs in class_subjects if not cs.teacher_id)
    edt_no_slot    = (sum(1 for cs in class_subjects if cs.id not in scheduled_cs_ids)
                      if year else 0)

    return {
        'school_class':    school_class,
        'edt_year':        year,
        'day_columns':     day_columns,
        'edt_breaks':      breaks,
        'edt_subjects':    class_subjects,
        'edt_durations':   durations,
        'edt_no_teacher':  edt_no_teacher,
        'edt_no_slot':     edt_no_slot,
        'hour_marks':      hour_marks,
        'grid_height':     _GRID_END_MIN - _GRID_START_MIN,
        'weekday_choices': days,
    }


def _render_timetable(request, school, school_class):
    from django.template.loader import render_to_string
    return render_to_string('schools/partials/class_timetable.html',
                            _timetable_ctx(school, school_class), request=request)


def _edt_toast(message, msg_type='error', status=422):
    resp = HttpResponse(status=status)
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': message, 'type': msg_type}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def slot_save(request, class_id, slot_id=None):
    """Crée / édite un créneau. Les conflits (classe, prof, salle) sont refusés par
    CourseSlot.clean() → toast 422 avec le message précis (« M. X déjà en 5B… »)."""
    from .models import CourseSlot, ClassSubject
    from .periods import active_year_for
    from django.core.exceptions import ValidationError as DjValidationError

    school       = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)
    year = active_year_for(school)
    if year is None:
        return _edt_toast("Aucune année scolaire — créez-en une d'abord.")

    instance = None
    if slot_id is not None:
        instance = get_object_or_404(
            CourseSlot, id=slot_id, school_year__school=school,
            class_subject__school_class=school_class,
        )

    cs = get_object_or_404(
        ClassSubject, id=request.POST.get('class_subject'),
        school_class=school_class, is_active=True,
    )
    slot = instance or CourseSlot(school_year=year)
    slot.class_subject = cs
    slot.day        = int(request.POST.get('day', 0))
    slot.room       = (request.POST.get('room') or '').strip()
    try:
        from datetime import time as _time
        sh, sm = (request.POST.get('start_time') or '').split(':')
        eh, em = (request.POST.get('end_time') or '').split(':')
        slot.start_time, slot.end_time = _time(int(sh), int(sm)), _time(int(eh), int(em))
    except (ValueError, AttributeError):
        return _edt_toast('Heures invalides.')

    try:
        slot.full_clean()
    except DjValidationError as e:
        return _edt_toast(' '.join(e.messages))
    slot.save()

    resp = HttpResponse(_render_timetable(request, school, school_class))
    resp['HX-Trigger'] = json.dumps({'showToast': {
        'message': 'Créneau enregistré.' if instance is None else 'Créneau modifié.',
        'type': 'success'}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def slot_delete(request, class_id, slot_id):
    from .models import CourseSlot
    school       = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)
    slot = get_object_or_404(
        CourseSlot, id=slot_id, school_year__school=school,
        class_subject__school_class=school_class,
    )
    slot.delete()   # sans risque : l'historique de paie ne référence JAMAIS un créneau
    resp = HttpResponse(_render_timetable(request, school, school_class))
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Créneau supprimé.', 'type': 'info'}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def break_save(request, class_id):
    """Ajoute une pause nommée (NIVEAU ÉCOLE — visible sur toutes les classes)."""
    from .models import SchoolBreak
    from django.core.exceptions import ValidationError as DjValidationError

    school       = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)

    label = (request.POST.get('label') or '').strip()
    if not label:
        return _edt_toast('Donnez un nom à la pause (ex. Récréation).')
    day_raw = request.POST.get('day', '')
    try:
        from datetime import time as _time
        sh, sm = (request.POST.get('start_time') or '').split(':')
        eh, em = (request.POST.get('end_time') or '').split(':')
        brk = SchoolBreak(
            school=school, label=label,
            day=int(day_raw) if day_raw != '' else None,
            start_time=_time(int(sh), int(sm)), end_time=_time(int(eh), int(em)),
        )
        brk.full_clean()
    except (ValueError, AttributeError):
        return _edt_toast('Heures invalides.')
    except DjValidationError as e:
        return _edt_toast(' '.join(e.messages))
    brk.save()

    resp = HttpResponse(_render_timetable(request, school, school_class))
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Pause ajoutée.', 'type': 'success'}})
    return resp


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def break_delete(request, class_id, break_id):
    from .models import SchoolBreak
    school       = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)
    get_object_or_404(SchoolBreak, id=break_id, school=school).delete()
    resp = HttpResponse(_render_timetable(request, school, school_class))
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': 'Pause supprimée.', 'type': 'info'}})
    return resp


# ── Impression / consultation : emploi du temps d'une CLASSE ou d'un PROF ──────
# Une seule page autonome (toolbar Retour/Imprimer masquée à l'impression, A4
# paysage, N&B friendly) sert de vue lecture seule ET de document mural.

def _print_timetable_ctx(school, *, school_class=None, teacher=None):
    """Colonnes positionnées (1 px = 1 min) pour la page d'impression.
    Bloc classe → « Matière / prof » ; bloc prof → « Classe / matière ».
    Même règle dimanche que la grille : colonne seulement si contenu dominical."""
    from .models import CourseSlot, SchoolBreak, Weekday
    from .periods import active_year_for

    year = active_year_for(school)
    days = list(Weekday.choices)

    qs = CourseSlot.objects.none()
    if year:
        qs = (CourseSlot.objects.filter(school_year=year)
              .select_related('class_subject__subject', 'class_subject__teacher',
                              'class_subject__school_class')
              .order_by('day', 'start_time'))
        if school_class is not None:
            qs = qs.filter(class_subject__school_class=school_class)
        else:
            qs = qs.filter(class_subject__school_class__school=school,
                           class_subject__teacher=teacher)

    slots_by_day = {d: [] for d, _l in days}
    for s in qs:
        top    = max(_min_of(s.start_time) - _GRID_START_MIN, 0)
        height = max(_min_of(s.end_time) - max(_min_of(s.start_time), _GRID_START_MIN), 18)
        cs = s.class_subject
        if school_class is not None:
            line1, line2 = cs.subject.name, (cs.teacher.full_name if cs.teacher else '')
        else:
            line1, line2 = cs.school_class.name, cs.subject.name
        hours = f'{s.start_time:%H:%M}–{s.end_time:%H:%M}'
        if s.room:
            hours += f' · {s.room}'
        slots_by_day[s.day].append({'top': top, 'height': height,
                                    'line1': line1, 'line2': line2, 'hours': hours})

    breaks = list(SchoolBreak.objects.filter(school=school))
    sunday_visible = bool(slots_by_day.get(6)) or any(b.day == 6 for b in breaks)
    visible_days = [d for d, _l in days if d != 6 or sunday_visible]

    breaks_by_day = {d: [] for d, _l in days}
    for b in breaks:
        top    = max(_min_of(b.start_time) - _GRID_START_MIN, 0)
        height = max(_min_of(b.end_time) - max(_min_of(b.start_time), _GRID_START_MIN), 12)
        for d in ([b.day] if b.day is not None else visible_days):
            breaks_by_day[d].append({'label': b.label, 'top': top, 'height': height})

    return {
        'columns': [{'label': label, 'slots': slots_by_day[d], 'breaks': breaks_by_day[d]}
                    for d, label in days if d in visible_days],
        'hour_marks':  [{'label': f'{h}h', 'top': h * 60 - _GRID_START_MIN} for h in range(8, 18)],
        'grid_height': _GRID_END_MIN - _GRID_START_MIN,
        'edt_year':    year,
        'school':      school,
    }


@login_required
@director_or_staff_required
def class_timetable_print(request, class_id):
    """Emploi du temps de la CLASSE — page autonome consultable et imprimable."""
    school       = get_school(request)
    school_class = get_object_or_404(SchoolClass, id=class_id, school=school, is_active=True)
    ctx = _print_timetable_ctx(school, school_class=school_class)
    ctx.update({
        'title':    f'Emploi du temps — {school_class.name}',
        'subtitle': school_class.get_level_display(),
        'back_url': reverse('schools:class-detail', args=[school_class.id]),
    })
    return render(request, 'schools/timetable_print.html', ctx)


@login_required
@director_or_staff_required
def teacher_timetable_print(request, user_id):
    """Emploi du temps d'un PROF (dérivé de ses cours) — consultable et imprimable."""
    from apps.accounts.models import User
    school  = get_school(request)
    teacher = get_object_or_404(
        User, id=user_id, memberships__school=school, memberships__is_active=True,
    )
    ctx = _print_timetable_ctx(school, teacher=teacher)
    ctx.update({
        'title':    f'Emploi du temps — {teacher.full_name}',
        'subtitle': 'Enseignant',
        'back_url': reverse('team:detail', args=[teacher.id]),
    })
    return render(request, 'schools/timetable_print.html', ctx)
