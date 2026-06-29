"""
Vues module bulletins — /bulletins/
Étape 3/3 bulletins.
"""
import io
import json
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction

from apps.accounts.models import UserRole
from django.db.models import Avg, Count, Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from apps.core.mixins import get_school, director_or_staff_required
from apps.students.models import Student

from .models import (
    Bulletin, BulletinConfig, BulletinLine, ClassSubject,
    Note, Period, SchoolClass, SchoolYear, NoteSystem, AppreciationScale,
)
from .services.bulletin_calculator import BulletinCalculator
from .services.bulletin_pdf import (
    generate_bulletin_pdf, generate_class_pdf, save_bulletin_pdf,
)


calculator = BulletinCalculator()

LEVEL_BADGE = {
    'prescolaire':    ('Préscolaire',       'bg-purple-100 text-purple-700 border-purple-200'),
    'fondamental_1':  ('Fond. 1er Cycle',   'bg-blue-100 text-blue-700 border-blue-200'),
    'fondamental_2':  ('Fond. 2ème Cycle',  'bg-indigo-100 text-indigo-700 border-indigo-200'),
    'secondaire_gen': ('Secondaire Gén.',   'bg-green-100 text-green-700 border-green-200'),
    'secondaire_pro': ('Secondaire Pro',    'bg-teal-100 text-teal-700 border-teal-200'),
    'superieur':      ('Supérieur',         'bg-orange-100 text-orange-700 border-orange-200'),
}


# ─────────────────────────────────────────────────────────────
# Vue 1 : Page principale
# ─────────────────────────────────────────────────────────────

@login_required
def bulletins_main(request):
    """
    Page principale /bulletins/ avec 3 onglets.
    Paramètres GET : year, period, class, tab
    """
    school = get_school(request)

    # Années scolaires
    years = list(school.school_years.order_by('-start_date'))
    active_year = None
    year_id = request.GET.get('year')
    if year_id:
        active_year = next((y for y in years if str(y.pk) == year_id), None)
    if not active_year:
        active_year = (
            school.school_years.filter(is_active=True).first()
            or (years[0] if years else None)
        )

    # Périodes
    periods = []
    active_period = None
    if active_year:
        periods = list(active_year.periods.all())
        period_id = request.GET.get('period')
        if period_id:
            active_period = next((p for p in periods if str(p.pk) == period_id), None)
        if not active_period:
            active_period = periods[0] if periods else None

    # Classes
    classes = list(
        school.classes.filter(is_active=True)
        .order_by('level', 'name')
    )
    active_class = None
    class_id = request.GET.get('class')
    if class_id:
        active_class = next((c for c in classes if str(c.pk) == class_id), None)

    # Onglet actif
    active_tab = request.GET.get('tab', 'health')

    # Données pour l'onglet par défaut
    context = {
        'school':          school,
        'years':           years,
        'active_year':     active_year,
        'periods':         periods,
        'active_period':   active_period,
        'classes':         classes,
        'active_class':    active_class,
        'active_tab':      active_tab,
        'active_section':  'bulletins',
        'can_generate':    request.user.role in (UserRole.DIRECTOR, UserRole.STAFF) or request.user.is_superuser,
    }

    # Stats globales si classe + période sélectionnées
    if active_class and active_period:
        stats = _get_class_stats(active_class, active_period, school)
        context.update(stats)
        context['generated_count'] = len(stats['bulletins'])
        context['total_count'] = stats['student_count']
        context['pending_count'] = stats['student_count'] - len(stats['bulletins'])

    if active_class:
        context['subject_count'] = ClassSubject.objects.filter(
            school_class=active_class, is_active=True,
        ).count()
        context['active_class_badge'] = LEVEL_BADGE.get(
            active_class.level,
            ('', 'bg-gray-100 text-gray-600 border-gray-200'),
        )

    if active_class and active_period:
        gen   = context.get('generated_count', 0)
        total = context.get('total_count', 0)
        context['page_subtitle'] = f"{gen}/{total} bulletins · {active_period.name}"
    elif active_period:
        context['page_subtitle'] = active_period.name

    return render(request, 'bulletins/bulletins_main.html', context)


# ─────────────────────────────────────────────────────────────
# Vue 2-4 : Onglets HTMX
# ─────────────────────────────────────────────────────────────

@login_required
def health_tab(request):
    """Onglet Santé éducative — partial HTMX."""
    school = get_school(request)
    active_class, active_period = _get_class_and_period(request, school)
    if not active_class or not active_period:
        return HttpResponse('<p class="text-gray-400 text-sm">Sélectionnez une classe et une période.</p>')

    ctx = _get_class_stats(active_class, active_period, school)
    ctx['active_class'] = active_class
    ctx['active_period'] = active_period
    ctx['can_generate'] = request.user.role in (UserRole.DIRECTOR, UserRole.STAFF) or request.user.is_superuser

    return render(request, 'bulletins/partials/health_tab.html', ctx)


@login_required
def bulletins_tab(request):
    """Onglet Bulletins — partial HTMX."""
    school = get_school(request)
    active_class, active_period = _get_class_and_period(request, school)
    if not active_class or not active_period:
        return HttpResponse('<p class="text-gray-400 text-sm">Sélectionnez une classe et une période.</p>')

    students = list(
        Student.objects
        .filter(school_class=active_class, school=school, is_active=True)
        .order_by('full_name')
    )

    # Bulletins existants
    existing = {
        b.student_id: b
        for b in Bulletin.objects.filter(
            period=active_period,
            school_class=active_class,
            is_cancelled=False,
        ).select_related('student')
    }

    with_notes = _students_with_notes(active_period, active_class)
    rows = []
    for student in students:
        bul = existing.get(student.pk)
        rows.append({
            'student':    student,
            'bulletin':   bul,
            'has_notes':  student.pk in with_notes,
        })

    generated_count = len(existing)
    total_count     = len(students)
    stale_count     = _annotate_stale(rows, active_class, active_period)
    return render(request, 'bulletins/partials/bulletins_tab.html', {
        'rows':            rows,
        'school_class':    active_class,
        'period':          active_period,
        'can_generate':    request.user.role in (UserRole.DIRECTOR, UserRole.STAFF) or request.user.is_superuser,
        'generated_count': generated_count,
        'total_count':     total_count,
        'pending_count':   total_count - generated_count,
        'generated_pct':   int(generated_count / total_count * 100) if total_count > 0 else 0,
        'stale_count':     stale_count,
    })


@login_required
def rankings_tab(request):
    """Onglet Classements — partial HTMX."""
    school = get_school(request)
    active_class, active_period = _get_class_and_period(request, school)
    if not active_class or not active_period:
        return HttpResponse('<p class="text-gray-400 text-sm">Sélectionnez une classe et une période.</p>')

    bulletins = list(
        Bulletin.objects.filter(
            period=active_period,
            school_class=active_class,
            is_cancelled=False,
            general_average__isnull=False,
        )
        .select_related('student')
        .order_by('-general_average')
    )

    averages = [b.general_average for b in bulletins]
    class_average = round(sum(averages) / len(averages), 2) if averages else None

    return render(request, 'bulletins/partials/rankings_tab.html', {
        'bulletins':     bulletins,
        'school_class':  active_class,
        'period':        active_period,
        'class_average': class_average,
        'first_average': averages[0] if averages else None,
        'last_average':  averages[-1] if averages else None,
        'can_generate':  request.user.role in (UserRole.DIRECTOR, UserRole.STAFF) or request.user.is_superuser,
    })


# ─────────────────────────────────────────────────────────────
# Vue 5-6 : Génération
# ─────────────────────────────────────────────────────────────

@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def generate_class_bulletins(request, class_id, period_id):
    """Génère tous les bulletins d'une classe."""
    school = get_school(request)

    school_class = get_object_or_404(
        school.classes.filter(is_active=True), pk=class_id,
    )
    period = get_object_or_404(
        Period, pk=period_id, school_year__school=school,
    )

    # Conserver l'état publié : la régénération recrée des bulletins non publiés,
    # sinon « Régénérer » ferait perdre l'accès aux parents.
    published_ids = set(
        Bulletin.objects.filter(
            school_class=school_class, period=period, is_published=True, is_cancelled=False,
        ).values_list('student_id', flat=True)
    )
    try:
        bulletins = calculator.generate_class_bulletins(
            school_class, period, request.user,
        )
    except Exception as e:
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Erreur de génération : {e}', 'type': 'error'},
        })
        return resp

    if published_ids:
        from django.utils import timezone
        Bulletin.objects.filter(
            school_class=school_class, period=period,
            student_id__in=published_ids, is_cancelled=False,
        ).update(is_published=True, published_at=timezone.now())

    # Re-rendre l'onglet bulletins avec les nouveaux bulletins
    students = list(
        Student.objects
        .filter(school_class=school_class, school=school, is_active=True)
        .order_by('full_name')
    )
    existing = {
        b.student_id: b
        for b in Bulletin.objects.filter(
            period=period,
            school_class=school_class,
            is_cancelled=False,
        ).select_related('student')
    }
    with_notes = _students_with_notes(period, school_class)
    rows = []
    for student in students:
        bul = existing.get(student.pk)
        rows.append({
            'student':    student,
            'bulletin':   bul,
            'has_notes':  student.pk in with_notes,
        })

    generated_count = len(existing)
    total_count     = len(students)
    stale_count     = _annotate_stale(rows, school_class, period)
    tab_html = render_to_string('bulletins/partials/bulletins_tab.html', {
        'rows':            rows,
        'school_class':    school_class,
        'period':          period,
        'can_generate':    True,
        'generated_count': generated_count,
        'total_count':     total_count,
        'pending_count':   total_count - generated_count,
        'generated_pct':   int(generated_count / total_count * 100) if total_count > 0 else 0,
        'stale_count':     stale_count,
    }, request=request)
    # OOB : met à jour le badge compteur X/Y de l'en-tête (hors zone swappée).
    badge_html = render_to_string('bulletins/partials/_bulletins_badge.html', {
        'generated_count': generated_count, 'total_count': total_count, 'oob': True,
    }, request=request)
    response = HttpResponse(tab_html + badge_html)
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'{len(bulletins)} bulletin(s) généré(s) avec succès.',
            'type': 'success',
        },
        'bullets-generated': {
            'count':   len(bulletins),
            'classId': class_id,
            'periodId': period_id,
        }
    })
    return response


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def generate_student_bulletin(request, student_id, period_id):
    """Génère le bulletin d'un élève spécifique."""
    school = get_school(request)

    student = get_object_or_404(Student, pk=student_id, school=school, is_active=True)
    period = get_object_or_404(Period, pk=period_id, school_year__school=school)

    try:
        with transaction.atomic():
            bulletin = calculator.generate_bulletin(student, period, request.user)
            ranks = calculator.calculate_ranks(period, student.school_class)
            bulletin.rank = ranks.get(student.pk)
            bulletin.class_size = student.school_class.students.filter(is_active=True).count()
            bulletin.first_average = calculator.get_first_average(period, student.school_class)
            bulletin.save(update_fields=['rank', 'class_size', 'first_average'])
    except Exception as e:
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Erreur de génération : {e}', 'type': 'error'},
        })
        return resp

    sc = student.school_class
    total_count     = sc.students.filter(is_active=True).count()
    generated_count = Bulletin.objects.filter(
        period=period, school_class=sc, is_cancelled=False,
    ).count()

    row_html = render_to_string('bulletins/partials/bulletin_row.html', {
        'student':  student,
        'bulletin': bulletin,
    }, request=request)
    # OOB : badge compteur X/Y de l'en-tête.
    badge_html = render_to_string('bulletins/partials/_bulletins_badge.html', {
        'generated_count': generated_count, 'total_count': total_count, 'oob': True,
    }, request=request)
    response = HttpResponse(row_html + badge_html)
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Bulletin généré avec succès.', 'type': 'success'},
        'bullet-generated': {'studentId': student.pk},
    })
    return response


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def bulletin_publish(request, bulletin_id):
    """Publie un bulletin (is_published=True) et notifie les parents. Renvoie la ligne re-rendue."""
    from django.utils import timezone

    school = get_school(request)
    bulletin = get_object_or_404(
        Bulletin.objects.select_related('student', 'period', 'period__school_year'),
        pk=bulletin_id, student__school=school, is_cancelled=False,
    )

    if not bulletin.is_published:
        bulletin.is_published = True
        bulletin.published_at = timezone.now()
        bulletin.save(update_fields=['is_published', 'published_at'])

        # Notifier les parents (jamais bloquant)
        try:
            from apps.notifications.services import notify_guardians
            from apps.notifications.models import NotificationCategory
            notify_guardians(
                student=bulletin.student,
                category=NotificationCategory.BULLETIN,
                title=f'Bulletin disponible — {bulletin.period.name}',
                body=(
                    f'Le bulletin de {bulletin.student.full_name} '
                    f'pour {bulletin.period.name} est disponible.'
                ),
                url=reverse('parent:bulletins'),
                target=bulletin,
            )
        except Exception:
            pass

    response = render(request, 'bulletins/partials/bulletin_row.html', {
        'student':  bulletin.student,
        'bulletin': bulletin,
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'Bulletin de {bulletin.student.full_name} publié.',
            'type': 'success',
        },
    })
    return response


# ─────────────────────────────────────────────────────────────
# Vue 7-9 : Preview / Download
# ─────────────────────────────────────────────────────────────

@login_required
@director_or_staff_required
def bulletin_preview(request, bulletin_id):
    """Preview HTML d'un bulletin (modal)."""
    school = get_school(request)
    bulletin = get_object_or_404(
        Bulletin,
        pk=bulletin_id,
        student__school=school,
        is_cancelled=False,
    )
    lines = list(
        bulletin.lines.all()
        .select_related('class_subject__subject')
        .order_by('class_subject__order', 'class_subject__subject__name')
    )
    config = BulletinConfig.objects.filter(school=school).first()
    if not config:
        config = BulletinConfig.objects.create(school=school)

    return render(request, 'bulletins/bulletin_preview.html', {
        'bulletin': bulletin,
        'lines':    lines,
        'config':   config,
    })


@login_required
@director_or_staff_required
def bulletin_download(request, bulletin_id):
    """Téléchargement PDF d'un bulletin."""
    school = get_school(request)
    bulletin = get_object_or_404(
        Bulletin,
        pk=bulletin_id,
        student__school=school,
        is_cancelled=False,
    )

    try:
        pdf_bytes = generate_bulletin_pdf(bulletin)
    except Exception as e:
        messages.error(request, f'Impossible de générer le PDF : {e}')
        return redirect(
            f"{reverse('bulletins:main')}?year={bulletin.period.school_year_id}"
            f"&period={bulletin.period_id}&class={bulletin.school_class_id}&tab=bullets"
        )
    filename = (
        f'bulletin_{bulletin.student.full_name.replace(" ", "_")}_'
        f'{bulletin.period.name.replace(" ", "_")}.pdf'
    )

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@director_or_staff_required
def bulletin_view_pdf(request, bulletin_id):
    """Ouvre le PDF dans le navigateur (Content-Disposition: inline)."""
    school = get_school(request)
    bulletin = get_object_or_404(
        Bulletin,
        pk=bulletin_id,
        student__school=school,
        is_cancelled=False,
    )
    try:
        pdf_bytes = generate_bulletin_pdf(bulletin)
    except Exception as e:
        messages.error(request, f'Impossible de générer le PDF : {e}')
        return redirect(
            f"{reverse('bulletins:main')}?year={bulletin.period.school_year_id}"
            f"&period={bulletin.period_id}&class={bulletin.school_class_id}&tab=bullets"
        )
    filename = (
        f'bulletin_{bulletin.student.full_name.replace(" ", "_")}_'
        f'{bulletin.period.name.replace(" ", "_")}.pdf'
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
@director_or_staff_required
def bulletin_download_all(request, class_id, period_id):
    """Téléchargement ZIP de tous les bulletins d'une classe."""
    school = get_school(request)
    school_class = get_object_or_404(
        school.classes.filter(is_active=True), pk=class_id,
    )
    period = get_object_or_404(
        Period, pk=period_id, school_year__school=school,
    )

    bulletins = list(
        Bulletin.objects.filter(
            period=period,
            school_class=school_class,
            is_cancelled=False,
        ).select_related('student')
    )
    if not bulletins:
        return HttpResponse('Aucun bulletin généré.', status=404)

    try:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for bulletin in bulletins:
                pdf_bytes = generate_bulletin_pdf(bulletin)
                filename = (
                    f'{bulletin.student.full_name.replace(" ", "_")}_'
                    f'{bulletin.period.name.replace(" ", "_")}.pdf'
                )
                zf.writestr(filename, pdf_bytes)
    except Exception as e:
        messages.error(request, f'Impossible de générer le ZIP : {e}')
        return redirect(
            f"{reverse('bulletins:main')}?year={period.school_year_id}"
            f"&period={period.pk}&class={class_id}&tab=bullets"
        )

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = (
        f'attachment; filename="bulletins_{school_class.name}_'
        f'{period.name.replace(" ", "_")}.zip"'
    )
    return response


@login_required
@director_or_staff_required
def rankings_export(request, class_id, period_id):
    """Export Excel du classement d'une classe."""
    school = get_school(request)
    school_class = get_object_or_404(
        school.classes.filter(is_active=True), pk=class_id,
    )
    period = get_object_or_404(
        Period, pk=period_id, school_year__school=school,
    )

    bulletins = list(
        Bulletin.objects.filter(
            period=period,
            school_class=school_class,
            is_cancelled=False,
            general_average__isnull=False,
        )
        .select_related('student')
        .order_by('-general_average')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Classement {period.name}'[:31]  # Excel limite à 31 chars

    # En-tête
    headers = ['Rang', 'Nom et Prénom', 'Moyenne Générale (/20)', 'Appréciation']
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Données
    for rank, bul in enumerate(bulletins, start=1):
        row = [rank, bul.student.full_name, float(bul.general_average), bul.appreciation or '']
        ws.append(row)
        # Fond doré pour le 1er
        if rank == 1:
            gold_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            for cell in ws[ws.max_row]:
                cell.fill = gold_fill

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22

    filename = (
        f'classement_{school_class.name.replace(" ", "_")}_'
        f'{period.name.replace(" ", "_")}.xlsx'
    )
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────
# Helpers privés
# ─────────────────────────────────────────────────────────────

def _get_class_and_period(request, school):
    """Récupère la classe et la période depuis les paramètres GET."""
    class_id = request.GET.get('class')
    period_id = request.GET.get('period')

    active_class = None
    if class_id:
        active_class = school.classes.filter(pk=class_id, is_active=True).first()

    active_period = None
    if period_id:
        active_period = Period.objects.filter(
            pk=period_id,
            school_year__school=school,
        ).first()

    return active_class, active_period


def _student_has_notes(student, period, school_class):
    """Vérifie si un élève a au moins une note pour chaque matière de la classe."""
    cs_count = ClassSubject.objects.filter(
        school_class=school_class, is_active=True,
    ).count()
    if cs_count == 0:
        return False
    noted_count = Note.objects.filter(
        student=student,
        period=period,
        class_subject__school_class=school_class,
        is_cancelled=False,
    ).values('class_subject').distinct().count()
    return noted_count >= cs_count


def _students_with_notes(period, school_class):
    """
    Retourne un set de student_id ayant une note pour chaque matière active de la classe.
    2 requêtes SQL au lieu de 2×N.
    """
    cs_count = ClassSubject.objects.filter(
        school_class=school_class, is_active=True,
    ).count()
    if cs_count == 0:
        return set()

    rows = (
        Note.objects.filter(
            period=period,
            class_subject__school_class=school_class,
            is_cancelled=False,
        )
        .values('student_id')
        .annotate(distinct_subjects=Count('class_subject', distinct=True))
        .filter(distinct_subjects__gte=cs_count)
    )
    return {row['student_id'] for row in rows}


def _annotate_stale(rows, school_class, period):
    """Marque `row['is_stale']` = une note a été modifiée APRÈS la génération du bulletin
    → le bulletin (et son PDF/ce que voit le parent) est périmé. Renvoie le nb de périmés."""
    latest = dict(
        Note.objects.filter(class_subject__school_class=school_class, period=period)
        .values('student_id')
        .annotate(m=Max('modified_at'))
        .values_list('student_id', 'm')
    )
    stale = 0
    for row in rows:
        b = row.get('bulletin')
        last = latest.get(b.student_id) if b else None
        is_stale = bool(b) and last is not None and last > b.generated_at
        row['is_stale'] = is_stale
        if is_stale:
            stale += 1
    return stale


def _get_class_stats(school_class, period, school):
    """Calcule les stats pour l'onglet santé éducative."""
    students = list(
        Student.objects.filter(
            school_class=school_class, school=school, is_active=True,
        )
    )

    bulletins = {
        b.student_id: b
        for b in Bulletin.objects.filter(
            period=period,
            school_class=school_class,
            is_cancelled=False,
        ).select_related('student')
    }

    averages = [b.general_average for b in bulletins.values() if b.general_average is not None]

    # Moyennes par matière (agrégées sur tous les bulletins de la classe)
    subject_rows = (
        BulletinLine.objects
        .filter(
            bulletin__period=period,
            bulletin__school_class=school_class,
            bulletin__is_cancelled=False,
            final_average__isnull=False,
        )
        .values('class_subject__subject__name', 'class_subject__coefficient')
        .annotate(avg=Avg('final_average'))
        .order_by('-avg')
    )
    subject_stats = [
        {
            'name':        row['class_subject__subject__name'],
            'coefficient': row['class_subject__coefficient'],
            'avg':         round(float(row['avg']), 2),
            'pct':         round(float(row['avg']) / 20 * 100, 1),
        }
        for row in subject_rows
    ]

    return {
        'student_count':    len(students),
        'gen_avg':          round(sum(averages) / len(averages), 2) if averages else None,
        'success_rate':     (
            round(sum(1 for a in averages if a >= 10) / len(averages) * 100, 1)
            if averages else 0
        ),
        'admitted_count':   sum(1 for a in averages if a >= 10),
        'difficulty_count': sum(1 for a in averages if a < 10),
        'bulletins':        bulletins,
        'students':         students,
        'active_period':    period,
        'school_class':     school_class,
        'subject_stats':    subject_stats,
    }