import csv
import io
import json
import logging
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from apps.accounts.models import UserRole
from django.db import IntegrityError, transaction
from django.core.paginator import Paginator
from django.db.models import Count, DecimalField, F, Q, Subquery, OuterRef, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods

from apps.payments.models import Payment, PaymentMethod
from apps.schools.models import SchoolClass
from apps.schools.periods import periods_for_cycle, periods_for_student, resolve_active_period
from apps.core.mixins import get_school, director_or_staff_required
from apps.core.text import norm_name
from apps.dashboard.views import invalidate_dashboard_cache

from .forms import StudentCreateForm, StudentUpdateForm
from .models import (
    Student, StudentGuardian, ParentRelationship,
    StudentEnrollment, EnrollmentStatus,
)

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────



def _students_qs(school, status='all', class_id=None, no_parent=False):
    """
    Queryset de la liste élèves, annoté par le helper CENTRAL (nouveau modèle) :
    chaque élève porte fee_status / fee_due / fee_paid / fee_balance / has_fee_account.
    Les élèves SANS fiche ont fee_status='no_fee' (badge neutre, jamais rouge).
    Filtres COMBINABLES (axes indépendants) : statut paiement × classe × sans-parent.
    `status` ∈ {paid,partial,unpaid,no_fee} ; 'all' = pas de filtre statut.
    """
    from apps.finance.services import annotate_students_with_fees
    qs = (
        Student.objects
        .filter(school=school, is_active=True)
        .select_related('school_class')
        .order_by('last_name', 'first_name')
    )
    qs = annotate_students_with_fees(qs)
    if status in ('unpaid', 'partial', 'paid', 'no_fee'):
        qs = qs.filter(fee_status=status)
    if class_id:
        qs = qs.filter(school_class_id=class_id)
    if no_parent:
        qs = qs.filter(guardians__isnull=True)   # aucun responsable enregistré
    return qs


def compute_student_stats(school):
    """Stats de la liste — NOUVEAU modèle (lot 6bis-A). Solde global = Σ des soldes
    positifs des FICHES (élèves sans fiche exclus + comptés à part)."""
    from apps.finance.services import annotate_students_with_fees, count_without_account
    today = timezone.now().date()
    base = Student.objects.filter(school=school, is_active=True)
    unpaid_balance = (
        annotate_students_with_fees(base)
        .filter(has_fee_account=True, fee_balance__gt=0)
        .aggregate(total=Sum('fee_balance'))['total'] or 0
    )
    return {
        'total':               base.count(),
        'enrolled_today':      base.filter(enrolled_at__date=today).count(),
        'without_parent':      base.filter(guardians__isnull=True).count(),
        'unpaid_balance':      int(unpaid_balance),
        'without_fee_account': count_without_account(school),
    }


def _student_list_page(request, school):
    """Lit les filtres combinables (statut × classe × sans-parent × recherche) dans la
    query-string et renvoie le contexte PAGINÉ partagé par la page complète et le
    rafraîchissement HTMX (#student-list-area). 30 élèves / page."""
    status    = request.GET.get('status', 'all')
    class_id  = request.GET.get('class_id') or None
    no_parent = request.GET.get('no_parent') in ('1', 'true', 'on')
    query     = request.GET.get('q', '').strip()

    qs = _students_qs(school, status, class_id, no_parent)
    if query:
        # Recherche insensible casse + accents (normalisation Python). Échelle école.
        nq = norm_name(query)
        items = [
            s for s in qs
            if nq in norm_name(s.full_name)
            or nq in norm_name(s.school_class.name if s.school_class else '')
            or nq in norm_name(s.access_code)
        ]
    else:
        items = list(qs)

    page = Paginator(items, 30).get_page(request.GET.get('page'))
    return {
        'students':  page,
        'page_obj':  page,
        'status':    status,
        'class_id':  class_id,
        'no_parent': no_parent,
        'q':         query,
    }


# ── Vues principales ──────────────────────────────────────────────────────────

@login_required
def student_list(request):
    if request.user.role == UserRole.TEACHER:
        return redirect('teacher:dashboard')
    school = get_school(request)
    ctx = _student_list_page(request, school)

    classes = SchoolClass.objects.filter(school=school, is_active=True).order_by('level', 'name')
    # ── Données pour Alpine — passées comme OBJETS Python ──────────────────────
    # IMPORTANT : on laisse {% … json_script %} encoder UNE seule fois côté template.
    # Surtout PAS de json.dumps ici, sinon double encodage → JSON.parse renvoie une
    # chaîne et tout le composant Alpine casse (bug lot 4a).
    classes_data = [
        {'id': c.id, 'name': c.name, 'annual_fee': int(c.annual_fee), 'level': c.level}
        for c in classes
    ]
    # Catalogue de frais + gabarits pour le panneau enrichi (lot 4a).
    fees_data, schedule_data, default_template_count = _enrollment_catalog_data(school)

    ctx.update({
        'stats':         compute_student_stats(school),
        'form':          StudentCreateForm(school=school),
        'classes':       classes,
        'classes_data':  classes_data,
        'fees_data':     fees_data,
        'schedule_data': schedule_data,
        'default_template_count': default_template_count,
        'school':        school,
        'status_pills': [
            ('all', 'Tous'), ('paid', 'Soldé'), ('partial', 'Partiel'),
            ('unpaid', 'Impayé'), ('no_fee', 'Sans fiche'),
        ],
    })
    return render(request, 'students/student_list.html', ctx)


def _enrollment_catalog_data(school):
    """
    Données (objets Python, à encoder via json_script) pour le panneau enrichi (lot 4a) :
      - fees_data     : frais actifs hors scolarité (obligatoires + optionnels) + variantes ;
      - schedule_data : gabarits de tranches de l'école ;
      - default_template_count : nb de tranches du gabarit par défaut (1 par défaut).
    Robuste à une école NON configurée : renvoie des listes vides + 1 (repli annuel).
    """
    from apps.finance.models import FeeType, FeeCategory, PaymentScheduleTemplate

    fees = (
        FeeType.objects
        .filter(school=school, is_active=True)
        .exclude(category=FeeCategory.TUITION)
        .prefetch_related('variants')
        .order_by('is_mandatory', 'order', 'name')
    )
    fees_data = []
    for f in fees:
        fees_data.append({
            'id':              f.id,
            'name':            f.name,
            'category':        f.category,           # one_time | subscription
            'is_mandatory':    f.is_mandatory,
            'has_variants':    f.has_variants,
            'is_gender_based': f.is_gender_based,
            'default_amount':  int(f.default_amount) if f.default_amount is not None else None,
            'icon':            f.get_icon(),
            'variants': [
                {'id': v.id, 'label': v.label, 'amount': int(v.amount), 'gender_key': v.gender_key}
                for v in f.variants.all() if v.is_active
            ],
        })

    templates = list(
        PaymentScheduleTemplate.objects
        .filter(school=school, is_active=True)
        .order_by('installments_count')
    )
    schedule_data = [
        {'id': t.id, 'name': t.name, 'installments_count': t.installments_count,
         'is_default': t.is_default}
        for t in templates
    ]
    default_count = next((t.installments_count for t in templates if t.is_default), 1)

    # Objets Python (PAS de json.dumps) : json_script encodera une seule fois côté template.
    return fees_data, schedule_data, default_count


def _create_responsable_from_post(request, student, *, is_primary=False, prefix='responsable'):
    """Crée un responsable (StudentGuardian) depuis les champs POST `{prefix}_*`.

    L'INFO est toujours enregistrée (nom/téléphone/e-mail/lien). Un compte portail n'est
    créé/lié QUE si « accès portail » est coché et un téléphone fourni : compte existant
    par téléphone → lié ; sinon création d'un compte parent (mot de passe temporaire).
    Retourne le StudentGuardian, ou None si rien n'a été saisi.
    """
    from apps.students.models import StudentGuardian, ParentRelationship
    name  = request.POST.get(f'{prefix}_name', '').strip()
    phone = request.POST.get(f'{prefix}_phone', '').strip()
    rel   = request.POST.get(f'{prefix}_relationship', '').strip()
    email = request.POST.get(f'{prefix}_email', '').strip()
    portal = request.POST.get(f'{prefix}_portal') == 'on'
    if not (name or phone):
        return None

    guardian_user = None
    if portal and phone:
        from apps.accounts.models import User, UserRole
        from apps.accounts.team_forms import generate_temp_password
        guardian_user = User.objects.filter(phone_number=phone).first()
        if guardian_user is None:
            guardian_user = User.objects.create_user(
                phone_number=phone, password=generate_temp_password(),
                full_name=name or phone, role=UserRole.PARENT,
            )

    valid_rel = rel if rel in {c[0] for c in ParentRelationship.choices} else ''
    sg = StudentGuardian.objects.create(
        student=student, guardian=guardian_user,
        full_name=name, phone=phone, email=email,
        relationship=valid_rel, is_primary=is_primary,
    )
    if phone:
        logger.info('[SMS] Notification responsable à envoyer — élève : %s', student.full_name)
    return sg


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def student_create(request):
    school = get_school(request)

    # ── Garde-fou (lot 4a) : pas d'année active → pas d'inscription possible ────
    # La fiche financière s'accroche à l'enrollment de l'année active. Sans année
    # active, on bloque proprement (toast) au lieu d'inscrire un élève sans fiche.
    from apps.students.services import ensure_active_enrollment, has_active_year
    if not has_active_year(school):
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({'showToast': {
            'message': 'Aucune année scolaire active — configurez-en une avant d\'inscrire.',
            'type': 'error',
        }})
        return resp

    form = StudentCreateForm(request.POST, school=school)

    if form.is_valid():
        student = form.save(commit=False)
        student.school      = school
        student.tuition_fee = student.school_class.annual_fee
        student.save()  # gender inclus via le form (lot 4a)
        invalidate_dashboard_cache(school)

        # ── Fondation + fiche financière (lot 4a) ──────────────────────────────
        # 1) enrollment ACTIVE de l'année active (idempotent, contrainte unique lot 1).
        # 2) génération de la fiche avec les options cochées + le gabarit choisi.
        from apps.finance.services import build_fee_account
        from apps.finance.models import PaymentScheduleTemplate

        enrollment = ensure_active_enrollment(student)

        # Options optionnelles cochées : name="option_fees" (liste d'IDs FeeType) +
        # variant_<id> pour les frais à variantes (trajet de bus).
        fee_selections = []
        for fid in request.POST.getlist('option_fees'):
            try:
                fid_int = int(fid)
            except (TypeError, ValueError):
                continue
            vid = request.POST.get(f'variant_{fid}') or None
            try:
                vid_int = int(vid) if vid else None
            except (TypeError, ValueError):
                vid_int = None
            fee_selections.append({'fee_type_id': fid_int, 'variant_id': vid_int})

        # Gabarit de scolarité choisi (sinon défaut résolu dans build_fee_account).
        template = None
        tpl_id = request.POST.get('schedule_template')
        if tpl_id:
            template = PaymentScheduleTemplate.objects.filter(
                school=school, id=tpl_id, is_active=True,
            ).first()

        # On capture la fiche : ses tranches doivent exister AVANT d'allouer l'acompte.
        account = None
        if enrollment is not None:
            account = build_fee_account(enrollment, fee_selections=fee_selections, template=template)

        initial_amount = form.cleaned_data.get('initial_payment')
        if initial_amount and initial_amount > 0:
            payment = Payment.objects.create(
                student        = student,
                amount         = initial_amount,
                payment_method = form.cleaned_data.get('payment_method') or 'cash',
                collected_by   = request.user,
            )
            # ── Allocation de l'acompte (fix) ───────────────────────────────────
            # On réutilise STRICTEMENT le moteur du guichet (allocate_payment, lot 5),
            # cible par défaut = la SCOLARITÉ (FIFO sur ses tranches, cascade). Anti
            # sur-allocation géré par allocate_payment : si l'acompte dépasse le solde
            # scolarité, le surplus reste simplement non alloué (jamais de sur-allocation).
            # Cas limite (école sans annual_fee → pas de dette scolarité) : on alloue sur
            # la 1ère dette disponible, sinon on laisse non affecté sans planter.
            if account is not None:
                from apps.finance.services import allocate_payment
                from apps.finance.models import FeeDebtKind
                target = (account.debts.filter(kind=FeeDebtKind.TUITION).first()
                          or account.debts.first())
                if target is not None:
                    allocate_payment(payment, target)

        # ── Responsable principal (couche « responsable ») ─────────────────────
        # Info TOUJOURS ; compte portail SEULEMENT si « accès portail » coché. Un seul
        # responsable à l'inscription (les autres s'ajoutent sur la fiche élève).
        _create_responsable_from_post(request, student, is_primary=True)

        if request.htmx:
            ctx = _student_list_page(request, school)
            ctx['stats'] = compute_student_stats(school)
            ctx['success_message'] = f'{student.full_name} inscrit(e) — Code : {student.access_code}'
            response = render(request, 'students/partials/student_list_refresh.html', ctx)
            response['HX-Trigger'] = json.dumps({
                'close-panel': True,
                'showToast':   {'message': 'Élève inscrit avec succès.', 'type': 'success'},
            })
            return response

    elif request.htmx:
        return render(request, 'students/partials/student_form_fields.html', {'form': form})

    return redirect('students:list')


def _rail_summary(student, school, active_period):
    """KPIs « coup d'œil » du rail profil : statut/solde/% payé, prochaine échéance,
    moyenne (période active), absences 30j, observations non lues. Mono-élève → pas de N+1."""
    from apps.finance.services import student_fee_summary
    from apps.finance.models import StudentFeeAccount, FeeDebtKind
    from apps.schools.models import Note
    from apps.teachers.models import Attendance, StudentObservation
    today = timezone.now().date()

    summary = student_fee_summary(student)  # {due,paid,balance,status,has_overdue} | None
    pct_paid = 0
    if summary and summary['due']:
        pct_paid = int(round(summary['paid'] * 100 / summary['due']))

    # Prochaine échéance = tranche de scolarité non soldée la plus proche (échue ou à venir).
    next_inst = None
    acc = (StudentFeeAccount.objects
           .filter(enrollment__student=student, enrollment__status='active', enrollment__school=school)
           .prefetch_related('debts__installments__allocations').first())
    if acc:
        pend = [i for d in acc.debts.all() if d.kind == FeeDebtKind.TUITION
                for i in d.installments.all() if i.balance() > 0]
        pend.sort(key=lambda i: i.due_date)
        next_inst = pend[0] if pend else None

    moyenne = None
    if active_period:
        vals = list(Note.objects.filter(student=student, period=active_period, is_cancelled=False)
                    .values_list('value', flat=True))
        if vals:
            moyenne = round(sum(vals) / len(vals), 1)

    return {
        'fee_status':     summary['status'] if summary else 'no_fee',
        'fee_balance':    summary['balance'] if summary else None,
        'fee_paid':       summary['paid'] if summary else None,
        'fee_due':        summary['due'] if summary else None,
        'pct_paid':       pct_paid,
        'has_overdue':    summary['has_overdue'] if summary else False,
        'next_inst':      next_inst,
        'moyenne':        moyenne,
        'absences_count': Attendance.objects.filter(student=student, date__gte=today - timedelta(days=30)).count(),
        'unread_obs':     StudentObservation.objects.filter(student=student, school=school, is_private=False, is_read=False).count(),
    }


def _notes_by_subject(student, period):
    """Notes de la période regroupées par matière, avec moyenne par matière (onglet Scolarité)."""
    from apps.schools.models import Note
    if not period:
        return []
    notes = list(
        Note.objects.filter(student=student, period=period, is_cancelled=False)
        .select_related('class_subject', 'class_subject__subject')
        .order_by('class_subject__order', 'class_subject__subject__name', 'entered_at')
    )
    groups = {}
    for n in notes:
        groups.setdefault(n.class_subject.subject.name, []).append(n)
    out = []
    for subject, ns in groups.items():
        vals = [n.value for n in ns]
        out.append({'subject': subject, 'notes': ns,
                    'avg': round(sum(vals) / len(vals), 1) if vals else None})
    return out


@login_required
def student_detail(request, student_id):
    school = get_school(request)
    student = get_object_or_404(
        Student.objects.select_related('school_class').prefetch_related('payments'),
        id=student_id, school=school,
    )

    observations = []
    if request.role in (UserRole.DIRECTOR, UserRole.STAFF) or request.user.is_superuser:
        from apps.teachers.models import StudentObservation
        observations = list(
            StudentObservation.objects
            .filter(student=student, school=school, is_private=False)
            .select_related('teacher', 'read_by').order_by('-created_at')
        )

    guardians = student.guardians.select_related('guardian').order_by('-is_primary', 'created_at')

    from apps.teachers.models import Attendance
    from apps.schools.models import Period
    today = timezone.now().date()
    absences_recentes = list(
        Attendance.objects.filter(student=student, date__gte=today - timedelta(days=30))
        .select_related('teacher').order_by('-date')
    )

    active_period = (Period.objects
                     .filter(school_year__school=school, school_year__is_active=True)
                     .order_by('-is_notes_open', 'order').first())
    periods = list(Period.objects
                   .filter(school_year__school=school, school_year__is_active=True).order_by('order'))
    notes_by_subject = _notes_by_subject(student, active_period)
    recent_notes = sorted(
        (n for grp in notes_by_subject for n in grp['notes']),
        key=lambda n: n.entered_at, reverse=True,
    )[:5]

    notifs_parents = list(student.notifications.select_related('recipient').order_by('-created_at')[:20])

    # Fiche financière de l'enrollment ACTIVE — optionnelle (état neutre si absente).
    from apps.finance.models import StudentFeeAccount
    from apps.finance.services import timeline_families
    fee_account = (
        StudentFeeAccount.objects
        .filter(enrollment__student=student, enrollment__status='active', enrollment__school=school)
        .select_related('enrollment__school_year')
        .prefetch_related('debts__installments__allocations')
        .order_by('-enrollment__school_year__start_date').first()
    )
    fee_families = timeline_families(fee_account)

    ctx = {
        'student':           student,
        'school':            school,
        'observations':      observations,
        'guardians':         guardians,
        'absences_recentes': absences_recentes,
        'active_period':     active_period,
        'periods':           periods,
        'notes_by_subject':  notes_by_subject,
        'recent_notes':      recent_notes,
        'notifs_parents':    notifs_parents,
        'is_director':       request.role == UserRole.DIRECTOR or request.user.is_superuser,
        'fee_account':       fee_account,
        'fee_families':      fee_families,
        'tabs': [
            ('apercu', 'Aperçu'), ('finances', 'Finances'), ('scolarite', 'Scolarité'),
            ('vie-scolaire', 'Vie scolaire'), ('parents', 'Parents'),
        ],
    }
    ctx.update(_rail_summary(student, school, active_period))
    return render(request, 'students/student_detail.html', ctx)


@login_required
def student_detail_rail(request, student_id):
    """Rail identité + KPIs — rendu seul pour rafraîchissement HTMX après encaissement/édition."""
    school = get_school(request)
    student = get_object_or_404(
        Student.objects.select_related('school_class').prefetch_related('payments'),
        id=student_id, school=school,
    )
    from apps.schools.models import Period
    from apps.finance.models import StudentFeeAccount
    active_period = (Period.objects
                     .filter(school_year__school=school, school_year__is_active=True)
                     .order_by('-is_notes_open', 'order').first())
    fee_account = (StudentFeeAccount.objects
                   .filter(enrollment__student=student, enrollment__status='active',
                           enrollment__school=school).first())
    ctx = {'student': student, 'school': school, 'fee_account': fee_account,
           'is_director': request.role == UserRole.DIRECTOR or request.user.is_superuser}
    ctx.update(_rail_summary(student, school, active_period))
    return render(request, 'students/partials/student_rail.html', ctx)


@login_required
def student_notes_period(request, student_id):
    """Notes par matière pour une période donnée (sélecteur de l'onglet Scolarité)."""
    school = get_school(request)
    student = get_object_or_404(Student, id=student_id, school=school)
    # Périodes du cycle de l'élève (compositions / trimestres selon sa classe).
    periods = list(periods_for_student(student))
    pid = request.GET.get('period')
    period = next((p for p in periods if str(p.pk) == pid), None) if pid else None
    if not period:
        period = next((p for p in periods if p.is_notes_open), None) or (periods[0] if periods else None)
    return render(request, 'students/partials/notes_by_subject.html', {
        'student': student, 'active_period': period,
        'notes_by_subject': _notes_by_subject(student, period),
    })


@login_required
@director_or_staff_required
def observation_mark_read(request, student_id, obs_id):
    from apps.teachers.models import StudentObservation

    school = get_school(request)
    obs = get_object_or_404(
        StudentObservation.objects.select_related(
            'student', 'student__school_class', 'teacher', 'read_by',
        ),
        pk=obs_id,
        student_id=student_id,
        school=school,
        is_private=False,
    )
    if not obs.is_read:
        obs.is_read = True
        obs.read_at = timezone.now()
        obs.read_by = request.user
        obs.save(update_fields=['is_read', 'read_at', 'read_by'])

    # Inbox (Suivi) → ligne dense ; fiche élève → card complète.
    if request.GET.get('row'):
        return render(request, 'students/partials/obs_row.html', {'o': obs})
    return render(request, 'students/partials/obs_card.html', {
        'obs': obs, 'student': obs.student,
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def observation_share_parent(request, student_id, obs_id):
    """Toggle le partage d'une observation (non-privée) vers les parents + notifie."""
    from apps.teachers.models import StudentObservation
    from apps.notifications.services import notify_guardians, notify
    from apps.notifications.models import NotificationCategory

    school = get_school(request)
    student = get_object_or_404(Student, id=student_id, school=school)
    obs = get_object_or_404(
        StudentObservation,
        id=obs_id,
        student=student,
        is_private=False,
    )

    if obs.is_visible_to_parent:
        # Déjà partagé → retirer
        obs.is_visible_to_parent = False
        obs.save(update_fields=['is_visible_to_parent'])
        msg, notif_type = 'Observation retirée du portail parent.', 'info'
    else:
        # Partager + notifier les parents
        parent_message = request.POST.get('parent_message', '').strip()
        if parent_message:
            obs.parent_message = parent_message
        obs.is_visible_to_parent = True
        obs.save(update_fields=['is_visible_to_parent', 'parent_message'])
        message = obs.parent_message or obs.content[:100]
        notify_guardians(
            student=student,
            category=NotificationCategory.OBSERVATION,
            title=f"Message de l'école concernant {student.full_name}",
            body=message,
            url=reverse('parent:dashboard'),
            target=obs,
        )
        notify(
            recipient=obs.teacher,
            school=school,
            category=NotificationCategory.OBSERVATION,
            title="Votre observation a été partagée",
            body=f"Le directeur a partagé votre observation sur {student.full_name} avec les parents.",
            url=reverse('teacher:notifications'),
            target=obs,
        )
        msg, notif_type = 'Observation partagée avec les parents.', 'success'

    resp = render(request, 'students/partials/obs_share_button.html', {
        'obs': obs, 'student': student,
    })
    resp['HX-Trigger'] = json.dumps({
        'showToast': {'message': msg, 'type': notif_type},
    })
    return resp


@login_required
@director_or_staff_required
def student_update(request, student_id):
    school = get_school(request)
    student = get_object_or_404(
        Student.objects.select_related('school_class').prefetch_related('payments'),
        id=student_id, school=school,
    )

    if request.method == 'POST':
        form = StudentUpdateForm(request.POST, instance=student, school=school)
        if form.is_valid():
            student = form.save()
            # Re-fetch avec relations pour les méthodes financières
            student = get_object_or_404(
                Student.objects.select_related('school_class').prefetch_related('payments'),
                id=student_id, school=school,
            )
            if request.htmx:
                # Slide-over : pas de swap (204) → on ferme le panneau, rafraîchit le rail, toast.
                resp = HttpResponse(status=204)
                resp['HX-Trigger'] = json.dumps({
                    'close-edit-panel': True,
                    'refresh-rail':     True,
                    'showToast': {'message': 'Fiche élève mise à jour.', 'type': 'success'},
                })
                return resp
            return redirect('students:detail', student_id=student.id)

        # Erreurs de validation → on re-rend le formulaire dans le panneau.
        if request.htmx:
            return render(request, 'students/partials/student_edit_panel.html', {
                'student': student, 'form': form,
            })
        return render(request, 'students/student_detail.html', {
            'student': student, 'form': form, 'school': school,
        })

    # GET → formulaire d'édition pour le slide-over.
    form = StudentUpdateForm(instance=student, school=school)
    if request.htmx:
        return render(request, 'students/partials/student_edit_panel.html', {
            'student': student, 'form': form,
        })
    return redirect('students:detail', student_id=student.id)


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def student_withdraw(request, student_id):
    """Retire un élève des listes actives (transfert / abandon / fin d'année).

    Archive l'inscription (StudentEnrollment) puis Student.is_active=False.
    Les données (notes, paiements, bulletins) sont conservées (FK PROTECT).
    Action réservée au directeur.
    """
    school = get_school(request)
    if request.role != UserRole.DIRECTOR and not request.user.is_superuser:
        return HttpResponse(status=403)

    student = get_object_or_404(Student, id=student_id, school=school, is_active=True)

    status = request.POST.get('status')
    valid = (
        EnrollmentStatus.TRANSFERRED,
        EnrollmentStatus.GRADUATED,
        EnrollmentStatus.WITHDRAWN,
    )
    if status not in valid:
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({
            'showToast': {'message': 'Motif de retrait invalide.', 'type': 'error'},
        })
        return resp

    from apps.schools.models import SchoolYear
    active_year = SchoolYear.objects.filter(school=school, is_active=True).first()

    with transaction.atomic():
        StudentEnrollment.objects.create(
            student=student,
            school=school,
            school_class=student.school_class,
            school_year=active_year,
            status=status,
            enrolled_at=student.enrolled_at.date() if student.enrolled_at else None,
            ended_at=timezone.now().date(),
        )
        student.is_active = False
        student.save(update_fields=['is_active'])

    invalidate_dashboard_cache(school)

    messages.success(
        request,
        f'{student.full_name} retiré ({EnrollmentStatus(status).label}). '
        'Ses données sont conservées.',
    )
    resp = HttpResponse(status=204)
    resp['HX-Redirect'] = reverse('students:list')
    return resp


@login_required
def student_search(request):
    school = get_school(request)
    ctx = _student_list_page(request, school)
    return render(request, 'students/partials/student_table_body.html', ctx)


# ── Import Excel ──────────────────────────────────────────────────────────────

@login_required
def student_import_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Élèves'

    NUM_COLS = 8

    # ── Ligne 1 : instructions ────────────────────────────────────────
    ws.append([
        'OBLIGATOIRES : Nom, Prénom, Classe  |  OPTIONNELLES : Nom du responsable, '
        'Téléphone responsable, Date de naissance, Lien parenté, Genre (G/F)  |  '
        'Le matricule est attribué automatiquement. Les paiements se gèrent dans l\'application.'
    ])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    instr_cell = ws['A1']
    instr_cell.font      = Font(italic=True, color='555555')
    instr_cell.fill      = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
    instr_cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[1].height = 30

    # ── Ligne 2 : en-têtes ───────────────────────────────────────────
    headers = [
        'Nom *',
        'Prénom *',
        'Classe *',
        'Nom du responsable',
        'Téléphone responsable',
        'Date de naissance (JJ/MM/AAAA)',
        'Lien parenté (père/mère/tuteur)',
        'Genre (G/F)',   # lot 4b — optionnel, pilote la tenue auto
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for col_idx, cell in enumerate(ws[2], start=1):
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    # ── Lignes 3-4 : exemples ────────────────────────────────────────
    example_fill = PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid')
    for row_data in [
        ['Kouassi', 'Jean', 'CP1',    'Kouassi Ama',  '0700000002', '15/03/2015', 'mère', 'G'],
        ['Traoré',  'Awa',  '6ème A', 'Traoré Sékou', '0600000003', '20/07/2013', 'père', 'F'],
    ]:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.fill = example_fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_eleves.xlsx"'
    wb.save(response)
    return response


def _parse_student_rows(file_obj, filename, school):
    """Parse un fichier Excel/CSV et retourne (rows_valides, erreurs).

    Colonnes attendues (8) :
        Nom * | Prénom * | Classe * | Nom du responsable | Téléphone responsable
        | Date de naissance | Lien parenté | Genre
    """
    rows, errors = [], []

    class_map = {
        c.name.lower(): c
        for c in SchoolClass.objects.filter(school=school, is_active=True)
    }
    existing = set(
        Student.objects
        .filter(school=school, is_active=True)
        .values_list('full_name', 'school_class__name')
    )
    relationship_map = {
        'père': 'father', 'pere': 'father', 'papa': 'father', 'father': 'father',
        'mère': 'mother', 'mere': 'mother', 'mama': 'mother', 'mother': 'mother',
        'tuteur': 'guardian', 'tutrice': 'guardian', 'guardian': 'guardian',
    }
    # Genre (lot 4b) — tolérance large, normalisé vers les codes Gender du lot 1 ('M'/'F').
    # Cellule vide ou illisible → None (non bloquant : l'élève est créé sans genre).
    gender_map = {
        'f': 'F', 'fille': 'F', 'féminin': 'F', 'feminin': 'F', 'femme': 'F',
        'g': 'M', 'garçon': 'M', 'garcon': 'M', 'm': 'M',
        'masculin': 'M', 'homme': 'M',
    }

    try:
        if filename.lower().endswith('.csv'):
            content  = file_obj.read().decode('utf-8-sig')
            raw_rows = list(csv.reader(io.StringIO(content)))[1:]   # skip header row
            line_offset = 2
        else:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            # Détection automatique : cherche la ligne d'en-tête ("nom complet")
            # pour gérer les deux formats (avec ou sans ligne d'instructions)
            data_start = 2
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                first  = str(row[0] or '').strip().lower()
                second = str(row[1] or '').strip().lower() if len(row) > 1 else ''
                if 'nom' in first and 'prénom' in second:
                    data_start = i + 1
                    break
            raw_rows = [
                [str(cell.value).strip() if cell.value is not None else '' for cell in row]
                for row in ws.iter_rows(min_row=data_start)
            ]
            line_offset = data_start
    except Exception as exc:
        errors.append({'line': '—', 'name': '—', 'errors': [f'Impossible de lire le fichier : {exc}']})
        return rows, errors

    for line_num, raw in enumerate(raw_rows, start=line_offset):
        if not any(raw):
            continue
        cols = (raw + [''] * 8)[:8]
        last_raw, first_raw, class_raw, resp_name, resp_phone, dob_raw, rel_raw, gender_raw = [
            c.strip() for c in cols
        ]
        composed = f'{first_raw} {last_raw}'.strip()   # « Prénom Nom » pour l'affichage
        row_errors = []

        if not last_raw:
            row_errors.append('Nom manquant')
        if not first_raw:
            row_errors.append('Prénom manquant')

        school_class = class_map.get(class_raw.lower())
        if not school_class:
            row_errors.append(f'Classe « {class_raw} » introuvable')

        dob = None
        if dob_raw:
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    dob = datetime.strptime(dob_raw, fmt).date()
                    break
                except ValueError:
                    pass
            if not dob:
                row_errors.append(f'Date « {dob_raw} » non reconnue (attendu JJ/MM/AAAA)')

        parent_relationship = relationship_map.get(rel_raw.lower(), '') if rel_raw else ''
        # Genre : normalisé ou None. JAMAIS bloquant — pas une erreur de ligne.
        gender = gender_map.get(gender_raw.lower()) if gender_raw else None
        is_duplicate = bool(school_class and (composed, school_class.name) in existing)

        if row_errors:
            errors.append({'line': line_num, 'name': composed or '—', 'errors': row_errors})
        else:
            rows.append({
                'name':                composed,
                'last_name':           last_raw,
                'first_name':          first_raw,
                'class_id':            school_class.id,
                'class_name':          school_class.name,
                'dob':                 dob.isoformat() if dob else '',
                'resp_name':           resp_name,
                'resp_phone':          resp_phone,
                'parent_relationship': parent_relationship,
                'gender':              gender or '',     # '' = non renseigné
                'annual_fee':          int(school_class.annual_fee),
                'is_duplicate':        is_duplicate,
            })

    return rows, errors


@login_required
@require_http_methods(['POST'])
def student_import_preview(request):
    file_obj = request.FILES.get('import_file')
    if not file_obj:
        return HttpResponse('<p class="text-red-600 text-sm p-3">Aucun fichier sélectionné.</p>')

    school = get_school(request)
    rows, errors = _parse_student_rows(file_obj, file_obj.name, school)

    return render(request, 'students/partials/student_import_preview.html', {
        'rows':        rows,
        'parse_errors': errors,
        'duplicates':  [r['name'] for r in rows if r['is_duplicate']],
        'rows_json':   json.dumps(rows),
    })


def _batch_matricules(school, count, year=None):
    """`count` matricules séquentiels (AAAA-NNNN) pour l'année, à assigner avant bulk_create.

    Continue à partir du dernier numéro attribué à cette école pour l'année. Miroir en lot
    de generate_matricule() du modèle (bulk_create ne déclenche pas save()).
    """
    from django.utils import timezone
    year = year or timezone.now().year
    prefix = f'{year}-'
    max_seq = 0
    for m in (Student.objects
              .filter(school=school, matricule__startswith=prefix)
              .values_list('matricule', flat=True)):
        try:
            max_seq = max(max_seq, int(m.rsplit('-', 1)[-1]))
        except (ValueError, IndexError):
            continue
    return [f'{prefix}{max_seq + i:04d}' for i in range(1, count + 1)]


def _unique_access_codes(school, count):
    """Génère `count` codes à 6 chiffres uniques dans le lot ET absents en base pour cette école."""
    from .models import generate_student_access_code
    existing = set(
        Student.objects.filter(school=school)
        .values_list('access_code', flat=True)
    )
    codes, seen = [], set()
    attempts = 0
    while len(codes) < count:
        attempts += 1
        if attempts > count * 20:
            raise RuntimeError('Impossible de générer assez de codes uniques.')
        code = generate_student_access_code()
        if code not in existing and code not in seen:
            codes.append(code)
            seen.add(code)
    return codes


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def student_import_confirm(request):
    school = get_school(request)

    # ── Garde-fou (lot 4b) : pas d'année active → on bloque AVANT toute création ──
    # Option la plus sûre : ne rien créer plutôt que d'importer 1000 élèves sans fiche.
    from apps.students.services import has_active_year
    if not has_active_year(school):
        resp = HttpResponse(status=422)
        resp['HX-Trigger'] = json.dumps({'showToast': {
            'message': 'Configurez une année scolaire active avant d\'importer.',
            'type': 'error',
        }})
        return resp

    try:
        rows = json.loads(request.POST.get('rows_data', '[]'))
    except json.JSONDecodeError:
        return HttpResponse('<p class="text-red-600 text-sm p-3">Données invalides.</p>')

    class_cache = {
        c.id: c
        for c in SchoolClass.objects.filter(school=school, is_active=True)
    }

    students_to_create = []
    skipped = 0

    for row in rows:
        if row.get('is_duplicate'):
            skipped += 1
            continue
        sc = class_cache.get(row['class_id'])
        if not sc:
            skipped += 1
            continue

        # bulk_create contourne save() : on renseigne last_name/first_name (colonnes
        # séparées du modèle d'import) et le matricule (plus bas) explicitement.
        s = Student(
            school              = school,
            school_class        = sc,
            last_name           = row.get('last_name', ''),
            first_name          = row.get('first_name', ''),
            full_name           = row['name'],
            gender              = row.get('gender') or None,   # lot 4b — pilote la tenue auto
            tuition_fee         = sc.annual_fee,
        )
        if row.get('dob'):
            try:
                s.date_of_birth = date.fromisoformat(row['dob'])
            except ValueError:
                pass
        # Responsable (info seule, contact principal) créé après le bulk_create des élèves.
        s._resp = (row.get('resp_name', ''), row.get('resp_phone', ''), row.get('parent_relationship', ''))
        students_to_create.append(s)

    # Assigner des codes uniques (lot + base) avant bulk_create
    try:
        codes = _unique_access_codes(school, len(students_to_create))
    except RuntimeError as e:
        return HttpResponse(
            json.dumps({'showToast': {'message': str(e), 'type': 'error'}}),
            status=422,
            content_type='application/json',
        )
    for student, code in zip(students_to_create, codes):
        student.access_code = code

    # Matricules séquentiels pré-assignés (bulk_create contourne save() qui les génère).
    matricules = _batch_matricules(school, len(students_to_create))
    for student, mat in zip(students_to_create, matricules):
        student.matricule = mat

    try:
        created = Student.objects.bulk_create(students_to_create)
    except IntegrityError:
        # Fallback : save() un par un pour régénérer les codes en conflit
        created = []
        for student in students_to_create:
            while True:
                try:
                    student.pk = None
                    student.access_code = _unique_access_codes(school, 1)[0]
                    student.save()
                    created.append(student)
                    break
                except IntegrityError:
                    continue

    # ── Fondation + fiches (lot 4b) : enrollments + fiches en MASSE ─────────────
    # Les élèves importés sont NEUFS → on crée leurs enrollments en un seul bulk_create
    # (pas de get_or_create par élève : zéro N+1), puis build_fee_accounts_bulk génère
    # toutes les fiches en ~10 requêtes (catalogue/gabarit préchargés une fois).
    # Périmètre import (décision actée) : scolarité + frais obligatoires + tenue auto
    # par genre. Les options facultatives (bus/cantine) se cochent ensuite, élève par
    # élève, dans l'inscription individuelle. L'année active est garantie (garde-fou ci-dessus).
    from apps.students.models import StudentEnrollment, EnrollmentStatus
    from apps.schools.models import SchoolYear
    from apps.finance.services import build_fee_accounts_bulk

    active_year = SchoolYear.objects.get(school=school, is_active=True)
    today = timezone.now().date()
    enrollments = [
        StudentEnrollment(
            student=s, school=school, school_class=s.school_class,
            school_year=active_year, status=EnrollmentStatus.ACTIVE, enrolled_at=today,
        )
        for s in created
    ]
    StudentEnrollment.objects.bulk_create(enrollments)
    # Les objets `enrollments` portent déjà .student (avec gender) et .school_class
    # (avec annual_fee) en mémoire → build_fee_accounts_bulk n'émet aucune requête N+1.
    build_fee_accounts_bulk(enrollments)

    # ── Responsable principal par élève (info seule, sans compte) ───────────────
    from apps.students.models import StudentGuardian, ParentRelationship
    _valid_rel = {c[0] for c in ParentRelationship.choices}
    guardians = []
    for s in created:
        r_name, r_phone, r_rel = getattr(s, '_resp', ('', '', ''))
        if r_name or r_phone:
            guardians.append(StudentGuardian(
                student=s, guardian=None, full_name=r_name, phone=r_phone,
                relationship=r_rel if r_rel in _valid_rel else '', is_primary=True,
            ))
    if guardians:
        StudentGuardian.objects.bulk_create(guardians)

    if request.htmx:
        ctx = _student_list_page(request, school)
        ctx['stats'] = compute_student_stats(school)
        ctx['success_message'] = f'{len(created)} élève(s) importé(s), {skipped} ignoré(s).'
        response = render(request, 'students/partials/student_list_refresh.html', ctx)
        response['HX-Trigger'] = json.dumps({
            'close-import-modal': True,
            'showToast': {'message': f'{len(created)} élève(s) importé(s).', 'type': 'success'},
        })
        return response

    return redirect('students:list')


# ─────────────────────────────────────────────────────────────
# Phase D2 — Parents / Tuteurs (StudentGuardian)
# ─────────────────────────────────────────────────────────────

def _toast_error(message):
    resp = HttpResponse(status=422)
    resp['HX-Trigger'] = json.dumps({'showToast': {'message': message, 'type': 'error'}})
    return resp


def _render_guardian_section(request, student, *, toast=None, toast_type='success', close_panel=False):
    """Re-rend la section parents/tuteurs avec HX-Trigger (toast + fermeture panel)."""
    guardians = (
        student.guardians.select_related('guardian')
        .order_by('-is_primary', 'created_at')
    )
    resp = render(request, 'students/partials/guardian_section.html', {
        'student': student, 'guardians': guardians,
    })
    triggers = {}
    if toast:
        triggers['showToast'] = {'message': toast, 'type': toast_type}
    if close_panel:
        triggers['close-guardian-panel'] = True
    if triggers:
        resp['HX-Trigger'] = json.dumps(triggers)
    return resp


@login_required
@director_or_staff_required
def guardian_search(request, student_id):
    """GET ?phone= → cherche un compte parent. Partial HTMX (carte ou formulaire création)."""
    from apps.accounts.models import User
    from apps.accounts.team_forms import generate_temp_password

    school  = get_school(request)
    student = get_object_or_404(Student, id=student_id, school=school)
    phone   = request.GET.get('phone', '').strip()

    found, already_linked, blocked_role = None, False, None
    if phone:
        candidate = User.objects.filter(phone_number=phone).first()
        if candidate and candidate.role != UserRole.PARENT:
            blocked_role = candidate.get_role_display()
        elif candidate:
            found = candidate
            already_linked = student.guardians.filter(guardian=candidate).exists()

    return render(request, 'students/partials/guardian_search_result.html', {
        'student': student, 'phone': phone, 'searched': bool(phone),
        'found': found, 'already_linked': already_linked, 'blocked_role': blocked_role,
        'gen_password': generate_temp_password(),
        'relationships': ParentRelationship.choices,
    })


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def guardian_add(request, student_id):
    """Lier un parent existant (user_id) OU créer un compte parent (full_name+phone+password) + lier."""
    from apps.accounts.models import User
    from apps.accounts.team_forms import generate_temp_password

    school  = get_school(request)
    student = get_object_or_404(Student, id=student_id, school=school)

    relationship = request.POST.get('relationship', '')
    if relationship not in {c[0] for c in ParentRelationship.choices}:
        relationship = ''

    user_id = request.POST.get('user_id', '').strip()

    if user_id:
        parent = get_object_or_404(User, id=user_id, role=UserRole.PARENT)
    else:
        full_name = request.POST.get('full_name', '').strip()
        phone     = request.POST.get('phone', '').strip()
        if not full_name or not phone:
            return _toast_error('Nom et téléphone obligatoires.')
        if User.objects.filter(phone_number=phone).exists():
            return _toast_error('Ce numéro est déjà utilisé par un compte.')
        parent = User.objects.create_user(
            phone_number=phone,
            password=request.POST.get('password', '').strip() or generate_temp_password(),
            full_name=full_name, role=UserRole.PARENT,
        )

    is_first = not student.guardians.exists()
    link, created = StudentGuardian.objects.get_or_create(
        guardian=parent, student=student,
        defaults={'relationship': relationship, 'is_primary': is_first},
    )
    if not created:
        return _render_guardian_section(
            request, student,
            toast=f'{parent.full_name} est déjà lié à cet élève.', toast_type='info',
            close_panel=True,
        )
    return _render_guardian_section(
        request, student,
        toast=f'{parent.full_name} lié à l\'élève.', close_panel=True,
    )


@login_required
@director_or_staff_required
@require_http_methods(['POST'])
def guardian_remove(request, student_id, guardian_id):
    """Retire un lien parent (StudentGuardian.pk). Réassigne le contact principal si besoin."""
    school  = get_school(request)
    student = get_object_or_404(Student, id=student_id, school=school)
    link = get_object_or_404(StudentGuardian, id=guardian_id, student=student)

    name, was_primary = link.guardian.full_name, link.is_primary
    link.delete()
    if was_primary:
        nxt = student.guardians.order_by('created_at').first()
        if nxt:
            nxt.is_primary = True
            nxt.save(update_fields=['is_primary'])

    return _render_guardian_section(request, student, toast=f'{name} retiré.', toast_type='info')


# ─────────────────────────────────────────────────────────────
# Suivi global des élèves (admin) — /students/suivi/
# ─────────────────────────────────────────────────────────────

def _difficulty_flagged(school):
    """(abs_map, bul_map, obs_map, active_period, flagged_ids) — partagé entre la page et l'onglet."""
    from apps.teachers.models import Attendance, StudentObservation
    from apps.schools.models import Bulletin

    today = date.today()
    month_start = today.replace(day=1)

    active_year = school.school_years.filter(is_active=True).first()
    # Période active PAR cycle (compositions/trimestres) — l'établissement mélange
    # les deux. On agrège les moyennes sur toutes ces périodes à la fois.
    cycle_periods = []
    if active_year:
        cycles = set(school.classes.filter(is_active=True).values_list('level', flat=True))
        for cyc in cycles:
            p = resolve_active_period(periods_for_cycle(active_year, cyc))
            if p:
                cycle_periods.append(p)
    active_period = cycle_periods[0] if cycle_periods else None  # représentatif (entête)

    abs_map = dict(
        Attendance.objects.filter(school=school, status='absent', date__gte=month_start)
        .values('student_id').annotate(n=Count('id')).values_list('student_id', 'n')
    )
    # Moyenne académique = bulletin si généré (figé), sinon formatif (alerte précoce,
    # avant la composition). Fusion : formatif par défaut, bulletin non nul prioritaire.
    bul_map = {}
    if cycle_periods:
        from apps.schools.models import FormativeGrade
        from decimal import Decimal
        bulletin_raw = dict(
            Bulletin.objects.filter(student__school=school, is_cancelled=False, period__in=cycle_periods)
            .values_list('student_id', 'general_average')
        )
        acc = {}
        for sid, mx, val in (
            FormativeGrade.objects
            .filter(evaluation__period__in=cycle_periods, is_absent=False, value__isnull=False,
                    student__school=school)
            .values_list('student_id', 'evaluation__max_grade', 'value')
        ):
            acc.setdefault(sid, []).append(val / (mx or Decimal('20')) * 20)
        bul_map = {sid: round(sum(v) / len(v), 2) for sid, v in acc.items()}
        for sid, avg in bulletin_raw.items():
            if avg is not None:
                bul_map[sid] = avg
    obs_map = dict(
        StudentObservation.objects.filter(school=school, is_private=False, is_read=False)
        .values('student_id').annotate(n=Count('id')).values_list('student_id', 'n')
    )
    flagged_ids = (
        {sid for sid, n in abs_map.items() if n >= 3}
        | {sid for sid, avg in bul_map.items() if avg is not None and avg < 10}
        | set(obs_map)
    )
    return abs_map, bul_map, obs_map, active_period, flagged_ids


@login_required
@director_or_staff_required
def student_tracking(request):
    """Page suivi global (3 onglets HTMX). Stats résumé chargées d'emblée."""
    from apps.teachers.models import Attendance, StudentObservation
    from apps.notifications.models import Notification

    school = get_school(request)
    tab = request.GET.get('tab', 'absences')

    classes = school.classes.filter(is_active=True).order_by('level', 'name')

    today = timezone.now().date()
    month_start = today.replace(day=1)
    _, _, _, _, flagged_ids = _difficulty_flagged(school)
    stats = {
        'absences_today': Attendance.objects.filter(
            school=school, status='absent', date=today,
        ).count(),
        'obs_unread': StudentObservation.objects.filter(
            school=school, is_private=False, is_read=False,
        ).count(),
        'difficulty_count': len(flagged_ids),
        'notifs_sent_month': Notification.objects.filter(
            school=school, created_at__date__gte=month_start,
        ).count(),
    }

    return render(request, 'students/tracking.html', {
        'classes': classes,
        'tab': tab,
        'stats': stats,
        'school': school,
    })


@login_required
@director_or_staff_required
def tracking_absences(request):
    from apps.teachers.models import Attendance

    school   = get_school(request)
    class_id = request.GET.get('class')
    periode  = request.GET.get('periode', 'today')
    today    = date.today()

    qs = (
        Attendance.objects
        .filter(school=school, status__in=['absent', 'late'])
        .select_related('student', 'school_class', 'teacher')
        .order_by('-date')
    )
    if class_id:
        qs = qs.filter(school_class_id=class_id)
    if periode == 'today':
        qs = qs.filter(date=today)
    elif periode == 'week':
        qs = qs.filter(date__gte=today - timedelta(days=7))
    elif periode == 'month':
        qs = qs.filter(date__gte=today.replace(day=1))

    page = Paginator(qs, 50).get_page(request.GET.get('page', 1))
    classes = school.classes.filter(is_active=True).order_by('level', 'name')
    return render(request, 'students/partials/tracking_absences.html', {
        'absences': page, 'periode': periode, 'class_id': class_id, 'classes': classes,
    })


@login_required
@director_or_staff_required
def tracking_observations(request):
    from apps.teachers.models import StudentObservation

    school = get_school(request)
    filtre = request.GET.get('filtre', 'unread')   # défaut = les actionnables
    class_id = request.GET.get('class')

    qs = (
        StudentObservation.objects
        .filter(school=school, is_private=False)
        .select_related('student', 'student__school_class', 'teacher', 'read_by')
        .order_by('-created_at')
    )
    if filtre == 'unread':
        qs = qs.filter(is_read=False)
    elif filtre == 'shared':
        qs = qs.filter(is_visible_to_parent=True)
    if class_id:
        qs = qs.filter(student__school_class_id=class_id)

    page = Paginator(qs, 25).get_page(request.GET.get('page', 1))
    classes = school.classes.filter(is_active=True).order_by('level', 'name')
    return render(request, 'students/partials/tracking_observations.html', {
        'observations': page, 'filtre': filtre, 'class_id': class_id, 'classes': classes,
    })


@login_required
@director_or_staff_required
def tracking_difficulty(request):
    """Élèves signalés : absences>=3/mois OU moy<10 OU observations non lues. 4 requêtes."""
    school = get_school(request)
    niveau   = request.GET.get('niveau')     # critical | warning | watch
    class_id = request.GET.get('class')
    abs_map, bul_map, obs_map, active_period, flagged_ids = _difficulty_flagged(school)

    students = (
        Student.objects
        .filter(id__in=flagged_ids, school=school, is_active=True)
        .select_related('school_class')
    )
    if class_id:
        students = students.filter(school_class_id=class_id)

    results = []
    for s in students:
        abs_count = abs_map.get(s.pk, 0)
        avg = bul_map.get(s.pk)
        obs_count = obs_map.get(s.pk, 0)
        score = (
            2 * (abs_count >= 3)
            + 2 * (avg is not None and avg < 10)
            + 1 * (obs_count > 0)
        )
        results.append({
            'student': s, 'absences': abs_count, 'average': avg,
            'obs_unread': obs_count, 'score': score,
            'level': 'critical' if score >= 4 else ('warning' if score >= 2 else 'watch'),
        })
    if niveau in ('critical', 'warning', 'watch'):
        results = [r for r in results if r['level'] == niveau]
    results.sort(key=lambda x: -x['score'])

    page = Paginator(results, 30).get_page(request.GET.get('page', 1))
    classes = school.classes.filter(is_active=True).order_by('level', 'name')
    return render(request, 'students/partials/tracking_difficulty.html', {
        'results': page, 'period': active_period,
        'niveau': niveau, 'class_id': class_id, 'classes': classes,
    })
